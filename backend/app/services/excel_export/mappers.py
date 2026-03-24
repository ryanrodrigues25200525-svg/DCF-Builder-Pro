from __future__ import annotations

import re
from copy import copy
from datetime import date, datetime
from typing import Any

from openpyxl.cell.cell import MergedCell
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.workbook import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

SHEET_COVER = "Cover"
SHEET_OUTPUTS = "Outputs - Base"
SHEET_OUTPUTS_LEGACY = "Ouputs - Base"
SHEET_DCF_BASE = "DCF Model - Base (1)"
SHEET_DCF_BULL = "DCF Model - Bull (2)"
SHEET_DCF_BEAR = "DCF Model - Bear (3)"
SHEET_WACC = "WACC"
SHEET_COMPS = "Comps"
SHEET_ASSUMPTION_BREAKDOWN = "Assumption Breakdown"
SHEET_DATA_RECALCULATED = "Data Given (Recalculated)"
SHEET_DATA_ORIGINAL = "Original & Adjusted Data"

TEN_YEAR_COLUMNS = ["V", "W", "X", "Y", "Z", "AA", "AB", "AC", "AD", "AE"]
RECALC_COLUMNS = ["G", "H", "I", "J", "K", "L", "M", "N", "O", "P"]
DCF_TIMELINE_COLUMNS = ["H", "I", "J", "K", "L", "M", "N", "O", "P", "Q"]
WACC_LOOP_MODE_CURRENT_EQUITY = "current_equity"
WACC_LOOP_MODE_ITERATIVE = "iterative"
SCENARIO_BASE = "base"
SCENARIO_BULL = "bull"
SCENARIO_BEAR = "bear"
MAX_TERMINAL_GROWTH_RATE = 0.03
MIN_TERMINAL_WACC_SPREAD = 0.005
DCF_HELPER_CASH_CELL = "AA17"
DCF_HELPER_DEBT_CELL = "AA18"
DCF_HELPER_NON_OP_CELL = "AA19"
DCF_HELPER_CASH_CELL_ABS = "$AA$17"
DCF_HELPER_DEBT_CELL_ABS = "$AA$18"
DCF_HELPER_NON_OP_CELL_ABS = "$AA$19"


def apply_payload_to_workbook(workbook: Workbook, payload: dict[str, Any]) -> None:
    divisor = _resolve_amount_scale_divisor(payload)

    cover = _sheet(workbook, SHEET_COVER)
    outputs = _sheet(workbook, SHEET_OUTPUTS)
    if outputs.title != SHEET_OUTPUTS:
        outputs.title = SHEET_OUTPUTS
        _rewrite_formula_sheet_name_references(workbook, old_name=SHEET_OUTPUTS_LEGACY, new_name=SHEET_OUTPUTS)
    dcf_base = _sheet(workbook, SHEET_DCF_BASE)
    dcf_bull = _sheet(workbook, SHEET_DCF_BULL)
    dcf_bear = _sheet(workbook, SHEET_DCF_BEAR)
    dcf_bull.sheet_state = "visible"
    dcf_bear.sheet_state = "visible"
    wacc = _sheet(workbook, SHEET_WACC)
    comps_ws = _sheet(workbook, SHEET_COMPS)
    data_recalc = _sheet(workbook, SHEET_DATA_RECALCULATED)
    data_original = _sheet(workbook, SHEET_DATA_ORIGINAL)
    data_original.sheet_state = "hidden"
    if "Data ->" in workbook.sheetnames:
        workbook["Data ->"].sheet_state = "hidden"

    ticker = _payload_ticker(payload)
    company_name = _payload_company_name(payload)

    _map_cover_sheet(cover, payload, ticker, company_name)
    _map_company_labels(company_name, ticker, outputs, dcf_base, dcf_bull, dcf_bear, wacc, comps_ws)
    _map_currency_labels(outputs, dcf_base, dcf_bull, dcf_bear, wacc, comps_ws)

    _map_dcf_base_inputs(dcf_base, payload, divisor)
    assumptions = payload.get("assumptions", {})
    bear_nwc_multiplier = _to_float(assumptions.get("bearNwcMultiplier")) if isinstance(assumptions, dict) else None
    _sync_shared_scenario_inputs(dcf_bull, dcf_base, nwc_multiplier=1.0)
    _sync_shared_scenario_inputs(dcf_bear, dcf_base, nwc_multiplier=bear_nwc_multiplier or 1.0)
    _align_income_statement_labels(dcf_base, dcf_bull, dcf_bear, data_original, data_recalc)
    _normalize_dcf_waterfall_formulas(dcf_base, dcf_bull, dcf_bear)
    _harden_growth_rate_formulas(dcf_base, dcf_bull, dcf_bear)

    _map_wacc_inputs(wacc, payload)
    _apply_required_wacc_formulas(wacc, payload)
    _map_comps(comps_ws, payload, divisor)
    _harden_comps_ratio_formulas(comps_ws)
    _harden_wacc_peer_aggregate_formulas(wacc)

    timeline_years, historical_years = _build_timeline(payload)
    _map_year_headers(
        outputs,
        dcf_base,
        dcf_bull,
        dcf_bear,
        data_original,
        data_recalc,
        timeline_years,
        historical_years,
        payload,
    )
    _map_data_sheets(data_original, data_recalc, payload, divisor, timeline_years)
    _normalize_public_dcf_layout(outputs, dcf_base, dcf_bull, dcf_bear, payload, divisor)
    _sync_scenario_formula_backbone(dcf_base, dcf_bull, dcf_bear)
    _apply_capex_schedule_to_dcf(dcf_base, dcf_bull, dcf_bear, payload, timeline_years, divisor)
    _apply_scenario_snapshots_to_dcf(dcf_base, dcf_bull, dcf_bear, payload, timeline_years, divisor)
    _finalize_assumption_block_cleanup(dcf_base, dcf_bull, dcf_bear)
    _add_prior_actual_year_display_column(
        dcf_base,
        dcf_bull,
        dcf_bear,
        payload,
        timeline_years,
        historical_years,
        divisor,
    )
    _map_sensitivity_blocks(dcf_base, dcf_bull, dcf_bear, payload, divisor)
    _replace_template_placeholders(
        company_name=company_name,
        ticker=ticker,
        sheets=(cover, outputs, dcf_base, dcf_bull, dcf_bear, wacc, comps_ws, data_recalc, data_original),
    )
    # Final pass to keep output/scenario year headers as explicit FY text labels.
    _finalize_timeline_headers(outputs, dcf_base, dcf_bull, dcf_bear, timeline_years, historical_years)
    _reset_dcf_sheet_view_to_top(dcf_base, dcf_bull, dcf_bear)

    _remove_assumption_breakdown(workbook, cover)


def _sheet(workbook: Workbook, name: str) -> Worksheet:
    if name in workbook.sheetnames:
        return workbook[name]
    if name == SHEET_OUTPUTS and SHEET_OUTPUTS_LEGACY in workbook.sheetnames:
        return workbook[SHEET_OUTPUTS_LEGACY]
    raise KeyError(f"Missing required sheet: {name}")


def _to_float(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return number if number == number and number not in (float("inf"), float("-inf")) else None
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        try:
            number = float(stripped)
            return number if number == number and number not in (float("inf"), float("-inf")) else None
        except ValueError:
            return None
    return None


def _sanitize_wacc_rate(value: Any) -> float | None:
    parsed = _to_float(value)
    if parsed is None:
        return None
    return max(0.01, min(0.30, parsed))


def _sanitize_terminal_growth_rate(value: Any, *, reference_wacc: float | None = None) -> float | None:
    parsed = _to_float(value)
    if parsed is None:
        return None
    clamped = max(0.0, min(MAX_TERMINAL_GROWTH_RATE, parsed))
    if reference_wacc is not None:
        max_allowed = max(0.0, reference_wacc - MIN_TERMINAL_WACC_SPREAD)
        clamped = min(clamped, max_allowed)
    return clamped


def _rewrite_formula_sheet_name_references(workbook: Workbook, *, old_name: str, new_name: str) -> None:
    old_ref = f"'{old_name}'!"
    new_ref = f"'{new_name}'!"
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("=") and old_ref in value:
                    cell.value = value.replace(old_ref, new_ref)
                hyperlink = cell.hyperlink
                if hyperlink is not None and isinstance(hyperlink.location, str) and old_ref in hyperlink.location:
                    hyperlink.location = hyperlink.location.replace(old_ref, new_ref)


def _scenario_assumption_value(payload: dict[str, Any], scenario_name: str, key: str) -> Any:
    scenarios = payload.get("scenarios")
    if not isinstance(scenarios, dict):
        return None
    scenario = scenarios.get(scenario_name)
    if not isinstance(scenario, dict):
        return None
    assumptions = scenario.get("assumptions")
    if not isinstance(assumptions, dict):
        return None
    return assumptions.get(key)


def _set_comment(worksheet: Worksheet, cell_ref: str, text: str, *, author: str = "DCF Builder") -> None:
    cell = worksheet[cell_ref]
    if isinstance(cell, MergedCell):
        return
    if not text.strip():
        return
    cell.comment = Comment(text, author)


def _safe_set(worksheet: Worksheet, cell_ref: str, value: Any) -> None:
    _safe_set_with_options(worksheet, cell_ref, value, clear_if_none=False)


def _safe_set_or_clear(worksheet: Worksheet, cell_ref: str, value: Any) -> None:
    _safe_set_with_options(worksheet, cell_ref, value, clear_if_none=True)


def _safe_set_with_options(
    worksheet: Worksheet,
    cell_ref: str,
    value: Any,
    *,
    clear_if_none: bool,
) -> None:
    cell = worksheet[cell_ref]
    if isinstance(cell, MergedCell):
        return
    if cell.data_type == "f":
        return
    if value is None:
        if not clear_if_none:
            return
        cell.value = None
        cell.hyperlink = None
        return
    cell.value = value


def _force_set(worksheet: Worksheet, cell_ref: str, value: Any) -> None:
    cell = worksheet[cell_ref]
    if isinstance(cell, MergedCell):
        return
    cell.value = value


def resolve_wacc_loop_mode(payload: dict[str, Any]) -> str:
    assumptions = payload.get("assumptions", {})
    if not isinstance(assumptions, dict):
        return WACC_LOOP_MODE_CURRENT_EQUITY
    mode_raw = assumptions.get("waccLoopMode")
    if isinstance(mode_raw, str):
        mode = mode_raw.strip().lower()
        if mode in (WACC_LOOP_MODE_CURRENT_EQUITY, WACC_LOOP_MODE_ITERATIVE):
            return mode
    return WACC_LOOP_MODE_CURRENT_EQUITY


def _safe_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value.replace("Z", "+00:00")).date()
        except ValueError:
            return None
    return None


def _fiscal_year_end_date(raw: Any, fallback_year: int) -> date | None:
    parsed_date = _safe_date(raw)
    if parsed_date is not None:
        return parsed_date

    if not isinstance(raw, str):
        return None

    compact = raw.strip()
    numeric_match = re.match(r"^\s*(\d{1,2})[/-](\d{1,2})\s*$", compact)
    if numeric_match:
        month = int(numeric_match.group(1))
        day = int(numeric_match.group(2))
        try:
            return date(fallback_year, month, day)
        except ValueError:
            return None

    parts = compact.replace(",", " ").split()
    if len(parts) < 2:
        return None

    month_lookup = {
        "jan": 1,
        "january": 1,
        "feb": 2,
        "february": 2,
        "mar": 3,
        "march": 3,
        "apr": 4,
        "april": 4,
        "may": 5,
        "jun": 6,
        "june": 6,
        "jul": 7,
        "july": 7,
        "aug": 8,
        "august": 8,
        "sep": 9,
        "sept": 9,
        "september": 9,
        "oct": 10,
        "october": 10,
        "nov": 11,
        "november": 11,
        "dec": 12,
        "december": 12,
    }

    month = month_lookup.get(parts[0].lower())
    day = _to_float(parts[1])
    if month is None or day is None:
        return None

    try:
        return date(fallback_year, month, int(day))
    except ValueError:
        return None


def _safe_year_end_date(year: int, month: int, day: int) -> date:
    while day > 0:
        try:
            return date(year, month, day)
        except ValueError:
            day -= 1
    return date(year, month, 1)


def _series(record: dict[str, Any], keys: list[str]) -> list[float]:
    for key in keys:
        values = record.get(key)
        if isinstance(values, list):
            out: list[float] = []
            for item in values:
                parsed = _to_float(item)
                out.append(parsed if parsed is not None else 0.0)
            return out
    return []


def _resolve_amount_scale_divisor(payload: dict[str, Any]) -> float:
    company = payload.get("company", {})
    market = payload.get("market", {})
    historicals = payload.get("historicals", {})
    income = historicals.get("income", {}) if isinstance(historicals, dict) else {}

    unit_scale = str(company.get("unitsScale", "units")).lower()
    historical_revenue = _series(income, ["Revenue", "Total Revenue", "Sales"])

    forecast_values = []
    for forecast in payload.get("forecasts", []) or []:
        if not isinstance(forecast, dict):
            continue
        parsed = _to_float(forecast.get("revenue"))
        if parsed is not None:
            forecast_values.append(abs(parsed))

    sample_magnitude = max(
        [
            0.0,
            *[abs(v) for v in historical_revenue],
            *forecast_values,
            abs(_to_float(market.get("netDebt")) or 0.0),
            abs(_to_float(market.get("debt")) or 0.0),
        ]
    )

    if unit_scale == "billions":
        return 1_000_000_000.0
    if unit_scale == "millions":
        return 1_000_000.0 if sample_magnitude >= 1_000_000 else 1.0
    if unit_scale == "thousands":
        return 1_000.0 if sample_magnitude >= 1_000 else 1.0
    return 1.0


def _scale(value: float | None, divisor: float) -> float | None:
    if value is None:
        return None
    if divisor in (0, 1):
        return value
    return value / divisor


def _payload_ticker(payload: dict[str, Any]) -> str:
    ticker = payload.get("company", {}).get("ticker")
    if isinstance(ticker, str) and ticker.strip():
        return ticker.strip().upper()
    return "TICKER"


def _payload_company_name(payload: dict[str, Any]) -> str | None:
    company = payload.get("company", {})
    if isinstance(company, dict):
        name = company.get("name")
        if isinstance(name, str) and name.strip():
            return name.strip()
    ui_meta = payload.get("uiMeta", {})
    if isinstance(ui_meta, dict):
        name = ui_meta.get("companyName")
        if isinstance(name, str) and name.strip():
            return name.strip()
    return None


def _display_company_label(company_name: str | None, ticker: str) -> str:
    if company_name:
        return company_name
    return ticker


def _map_cover_sheet(cover: Worksheet, payload: dict[str, Any], ticker: str, company_name: str | None) -> None:
    company = payload.get("company", {})
    company = company if isinstance(company, dict) else {}
    ui_meta = payload.get("uiMeta", {})
    ui_meta = ui_meta if isinstance(ui_meta, dict) else {}

    author_name = ui_meta.get("author")
    author_email = ui_meta.get("authorEmail")
    _safe_set(cover, "C9", _display_company_label(company_name, ticker))
    _safe_set_or_clear(cover, "C10", company.get("industry") or company.get("sector") or "Technology")
    _safe_set(cover, "B12", "Scenario")
    _safe_set(cover, "C12", 1)
    _safe_set(cover, "C20", datetime.now().date())
    _safe_set_or_clear(cover, "C24", author_name if isinstance(author_name, str) and author_name.strip() else None)
    _safe_set_or_clear(cover, "C25", author_email if isinstance(author_email, str) and author_email.strip() else None)
    _apply_cover_scenario_validation(cover)


def _apply_cover_scenario_validation(cover: Worksheet) -> None:
    validation = DataValidation(type="list", formula1='"1,2,3"', allow_blank=False)
    validation.errorTitle = "Invalid Scenario"
    validation.error = "Select 1 (Base), 2 (Bull), or 3 (Bear)."
    validation.promptTitle = "Scenario"
    validation.prompt = "1=Base, 2=Bull, 3=Bear"
    cover.add_data_validation(validation)
    validation.add("C12")


def _map_company_labels(
    company_name: str | None,
    ticker: str,
    outputs: Worksheet,
    dcf_base: Worksheet,
    dcf_bull: Worksheet,
    dcf_bear: Worksheet,
    wacc: Worksheet,
    comps_ws: Worksheet,
) -> None:
    label = f"Company {_display_company_label(company_name, ticker)}"
    for sheet in (outputs, dcf_base, dcf_bull, dcf_bear, wacc, comps_ws):
        _safe_set(sheet, "B3", label)


def _map_currency_labels(
    outputs: Worksheet,
    dcf_base: Worksheet,
    dcf_bull: Worksheet,
    dcf_bear: Worksheet,
    wacc: Worksheet,
    comps_ws: Worksheet,
) -> None:
    usd_label = "All $ in USD millions unless otherwise stated"
    for sheet in (outputs, dcf_base, dcf_bull, dcf_bear, wacc, comps_ws):
        _safe_set(sheet, "B4", usd_label)


def _align_income_statement_labels(
    dcf_base: Worksheet,
    dcf_bull: Worksheet,
    dcf_bear: Worksheet,
    data_original: Worksheet,
    data_recalc: Worksheet,
) -> None:
    # Align template labels with website "Performance Matrix" naming.
    for sheet in (dcf_base, dcf_bull, dcf_bear):
        _safe_set(sheet, "B20", "Total Revenue")
        _safe_set(sheet, "B23", "Cost of Revenue")
        _safe_set(sheet, "B24", "Cost of Revenue")
        _safe_set(sheet, "B27", "Other Cost of Revenue")
        _safe_set(sheet, "B36", "Research & Development")
        _safe_set(sheet, "B39", "SG&A")
        _safe_set(sheet, "B42", "D&A (included in Operating)")
        _safe_set(sheet, "B45", "Other Operating Expenses")
        _safe_set(sheet, "B57", "Income Taxes")

    for sheet in (data_original, data_recalc):
        _safe_set(sheet, "B12", "Total Revenue")
        _safe_set(sheet, "B15", "Cost of Revenue")
        _safe_set(sheet, "B16", "Cost of Revenue")
        _safe_set(sheet, "B17", "Other Cost of Revenue")
        _safe_set(sheet, "B24", "Research & Development")
        _safe_set(sheet, "B25", "SG&A")
        _safe_set(sheet, "B26", "D&A (included in Operating)")
        _safe_set(sheet, "B27", "Other Operating Expenses")
        _safe_set(sheet, "B30", "EBIT")
        _safe_set(sheet, "B31", "EBIT Margin")


def _infer_revenue_growth_rate(forecasts: list[Any]) -> float | None:
    prev_revenue: float | None = None
    for item in forecasts:
        if not isinstance(item, dict):
            continue
        revenue = _to_float(item.get("revenue"))
        if revenue is None or revenue <= 0:
            continue
        if prev_revenue is not None and prev_revenue > 0:
            growth = (revenue / prev_revenue) - 1.0
            return max(-0.5, min(0.5, growth))
        prev_revenue = revenue
    return None


def _map_dcf_base_inputs(dcf_base: Worksheet, payload: dict[str, Any], divisor: float) -> None:
    assumptions = payload.get("assumptions", {})
    assumptions = assumptions if isinstance(assumptions, dict) else {}
    market = payload.get("market", {})
    market = market if isinstance(market, dict) else {}
    transaction = payload.get("transaction", {})
    transaction = transaction if isinstance(transaction, dict) else {}
    company = payload.get("company", {})
    company = company if isinstance(company, dict) else {}
    forecasts = payload.get("forecasts", []) or []
    key_metrics = (payload.get("uiMeta") or {}).get("keyMetrics") or {}

    tax_rate = _to_float(assumptions.get("taxRate"))
    da_pct = _to_float(assumptions.get("daPctRevenue"))
    if da_pct is None:
        da_pct = _to_float(assumptions.get("deaRatio"))
    terminal_assumptions = assumptions.get("terminal") or {}
    exit_multiple = _to_float(terminal_assumptions.get("exitMultiple"))
    wacc_assumptions = assumptions.get("wacc")
    wacc_assumptions = wacc_assumptions if isinstance(wacc_assumptions, dict) else {}
    base_wacc_assumption = _sanitize_wacc_rate(assumptions.get("waccRate") or wacc_assumptions.get("waccRate"))
    terminal_growth = _sanitize_terminal_growth_rate(
        terminal_assumptions.get("g"),
        reference_wacc=base_wacc_assumption,
    )
    revenue_growth = _to_float(assumptions.get("revenueGrowth"))
    if revenue_growth is None:
        revenue_growth = _to_float(assumptions.get("revenueGrowthRate"))
    if revenue_growth is None:
        revenue_growth = _infer_revenue_growth_rate(forecasts)
    if revenue_growth is None:
        revenue_growth = 0.06

    shares = _to_float(market.get("sharesDiluted")) or 0.0
    price = _to_float(market.get("currentPrice")) or 0.0
    market_cap_raw = _to_float(market.get("marketCap"))
    if market_cap_raw is None and shares > 0 and price > 0:
        market_cap_raw = shares * price

    debt_raw = _to_float(market.get("marketValueDebt"))
    if debt_raw is None:
        debt_raw = _to_float(market.get("debt")) or 0.0
    cash_raw = _to_float(market.get("cash")) or 0.0
    non_operating_assets_raw = _to_float(market.get("nonOperatingAssets")) or 0.0

    equity_raw = _to_float(key_metrics.get("equityValue"))
    if equity_raw is None:
        equity_raw = market_cap_raw

    net_debt_raw = _to_float(market.get("netDebt"))
    if net_debt_raw is None:
        net_debt_raw = debt_raw - cash_raw

    explicit_purchase_price = None
    if isinstance(transaction, dict):
        explicit_purchase_price = _to_float(transaction.get("purchasePrice"))
    if explicit_purchase_price is None:
        explicit_purchase_price = _to_float(payload.get("purchasePrice"))

    enterprise_raw = explicit_purchase_price if explicit_purchase_price is not None else _to_float(key_metrics.get("enterpriseValue"))
    if enterprise_raw is None and equity_raw is not None and net_debt_raw is not None:
        enterprise_raw = equity_raw + net_debt_raw

    _safe_set(dcf_base, "C9", _scale(enterprise_raw, divisor))
    _safe_set(dcf_base, "C11", _scale(equity_raw, divisor))
    _safe_set(dcf_base, "F16", _scale(net_debt_raw, divisor))
    _safe_set(dcf_base, "F17", _scale(cash_raw, divisor))
    _safe_set(dcf_base, "F18", _scale(debt_raw, divisor))
    _safe_set(dcf_base, "F19", _scale(non_operating_assets_raw, divisor))

    capex = None
    nwc_change = None

    first_forecast = forecasts[0] if forecasts and isinstance(forecasts[0], dict) else {}
    capex = _to_float(first_forecast.get("capex"))
    nwc_change = _to_float(first_forecast.get("nwcChange"))

    if capex is None:
        capex_abs = assumptions.get("capexAbsolute")
        if isinstance(capex_abs, list) and capex_abs:
            capex = _to_float(capex_abs[0])

    if capex is not None:
        capex = abs(capex)
    if nwc_change is not None:
        nwc_change = abs(nwc_change)

    _safe_set(dcf_base, "F9", _scale(capex, divisor))
    _safe_set(dcf_base, "F10", _scale(nwc_change, divisor))
    _safe_set(dcf_base, "F11", tax_rate)
    _safe_set(dcf_base, "F13", da_pct)
    _safe_set(dcf_base, "F14", revenue_growth)
    dcf_base["F14"].number_format = "0.0%"
    _safe_set(dcf_base, "C16", exit_multiple)
    _safe_set(dcf_base, "Q103", terminal_growth)

    _set_comment(dcf_base, "F9", "Source: Forecast capex (payload.forecasts[0].capex) or capexAbsolute fallback.")
    _set_comment(dcf_base, "F10", "Source: Forecast working-capital change (payload.forecasts[0].nwcChange).")
    _set_comment(dcf_base, "F11", "Source: Tax assumption from payload.assumptions.taxRate.")
    _set_comment(dcf_base, "F14", "Source: Revenue growth assumption; defaults to inferred forecast growth.")
    _set_comment(dcf_base, "C16", "Source: Terminal exit multiple from payload.assumptions.terminal.exitMultiple.")
    _set_comment(dcf_base, "F17", "Source: Market cash and equivalents.")
    _set_comment(dcf_base, "F18", "Source: Market debt / market value debt proxy.")
    _set_comment(dcf_base, "F19", "Source: Non-operating assets from market payload.")

    as_of = _safe_date(company.get("asOfDate"))
    if as_of is not None:
        _safe_set(dcf_base, "I9", as_of)

    fiscal_year = as_of.year if as_of is not None else datetime.now().year
    fiscal_end = _fiscal_year_end_date(company.get("fiscalYearEnd"), fiscal_year)
    if fiscal_end is None:
        fiscal_end = date(fiscal_year, 12, 31)
    _safe_set(dcf_base, "I11", fiscal_end)


def _sync_shared_scenario_inputs(scenario_sheet: Worksheet, dcf_base: Worksheet, *, nwc_multiplier: float) -> None:
    # Preserve scenario-specific formulas and apply only payload-driven shared assumptions.
    for cell in ("C9", "C11", "F9", "F11", "F13", "F14", "C16", "I9", "I11", "Q103"):
        scenario_sheet[cell].value = dcf_base[cell].value

    base_nwc_change = _to_float(dcf_base["F10"].value)
    if base_nwc_change is not None:
        scenario_sheet["F10"].value = base_nwc_change * max(0.0, nwc_multiplier)


def _map_wacc_inputs(wacc: Worksheet, payload: dict[str, Any]) -> None:
    assumptions = payload.get("assumptions", {})
    assumptions = assumptions if isinstance(assumptions, dict) else {}
    wacc_assumptions = assumptions.get("wacc", {})
    wacc_assumptions = wacc_assumptions if isinstance(wacc_assumptions, dict) else {}

    rf = _to_float(wacc_assumptions.get("rf"))
    if rf is None:
        rf = _to_float(assumptions.get("riskFreeRate"))
    if rf is None:
        rf = 0.044

    erp = _to_float(wacc_assumptions.get("erp"))
    if erp is None:
        erp = _to_float(assumptions.get("equityRiskPremium"))
    if erp is None:
        erp = 0.055

    illiquidity_discount = (
        _to_float(wacc_assumptions.get("illiquidityDiscount"))
        or _to_float(wacc_assumptions.get("liquidityDiscount"))
        or _to_float(assumptions.get("illiquidityDiscount"))
        or _to_float(assumptions.get("liquidityDiscount"))
    )

    size_premium = _to_float(wacc_assumptions.get("sizePremium"))
    if size_premium is None:
        size_premium = _to_float(assumptions.get("sizePremium"))
    if size_premium is None:
        size_premium = 0.0

    cost_of_debt = _to_float(wacc_assumptions.get("costOfDebt"))
    if cost_of_debt is None:
        cost_of_debt = _to_float(wacc_assumptions.get("currentDebtYield"))
    if cost_of_debt is None:
        cost_of_debt = _to_float(wacc_assumptions.get("debtYield"))
    if cost_of_debt is None:
        cost_of_debt = _to_float(assumptions.get("currentDebtYield"))
    if cost_of_debt is None:
        cost_of_debt = _to_float(assumptions.get("debtYield"))
    if cost_of_debt is None:
        credit_spread = _to_float(wacc_assumptions.get("creditSpread"))
        if credit_spread is None:
            credit_spread = _to_float(assumptions.get("creditSpread"))
        if credit_spread is not None:
            cost_of_debt = rf + credit_spread
    if cost_of_debt is None:
        cost_of_debt = _to_float(assumptions.get("costOfDebt"))
    if cost_of_debt is None:
        cost_of_debt = 0.051

    _safe_set(wacc, "D9", rf)
    _safe_set(wacc, "D10", erp)
    _safe_set(wacc, "D14", illiquidity_discount if illiquidity_discount is not None else 0.0)
    _safe_set(wacc, "D15", size_premium)
    _safe_set(wacc, "D19", cost_of_debt)
    _safe_set(wacc, "I7", "Debt ($B)")


def _apply_required_wacc_formulas(wacc: Worksheet, payload: dict[str, Any]) -> None:
    wacc_loop_mode = resolve_wacc_loop_mode(payload)
    assumptions = payload.get("assumptions", {})
    assumptions = assumptions if isinstance(assumptions, dict) else {}
    wacc_assumptions = assumptions.get("wacc")
    wacc_assumptions = wacc_assumptions if isinstance(wacc_assumptions, dict) else {}

    # Required fix: use beta from D11 directly.
    _force_set(wacc, "D11", "=H23")
    if wacc_loop_mode == WACC_LOOP_MODE_ITERATIVE:
        _force_set(wacc, "D23", "='DCF Model - Base (1)'!C12")
    else:
        _force_set(wacc, "D23", "='DCF Model - Base (1)'!C11")
    # Use market value of debt from helper cell, not net debt, for capital structure weights.
    _force_set(wacc, "D24", f"='DCF Model - Base (1)'!{DCF_HELPER_DEBT_CELL}")
    _force_set(wacc, "D27", "=IFERROR(D23/(D23+D24),0.85)")
    _force_set(wacc, "D28", "=IFERROR(D24/(D23+D24),0.15)")
    _force_set(wacc, "K23", "=IFERROR(IF(K17>0,K17,IF(D23>0,D24/D23,0.15)),0.15)")
    _force_set(wacc, "J23", "=IFERROR((D23+D24)/(1+K23),D23)")
    _force_set(wacc, "I23", "=IFERROR(D23+D24-J23,D24)")
    _force_set(wacc, "H23", "=M23*(1+(1-L23)*K23)")
    _force_set(wacc, "D16", "=D9+D11*(D10)+D14+D15")
    _force_set(wacc, "D20", "=IFERROR('DCF Model - Base (1)'!$F$11,0.21)")
    _force_set(wacc, "D21", "=IFERROR(D19*(1-D20),D19*0.79)")
    _force_set(wacc, "D31", "=IFERROR(J23/(I23+J23),0.85)")
    _force_set(wacc, "D32", "=IFERROR(I23/(I23+J23),0.15)")
    # Use current capital structure weights to avoid circular dependencies in optimization paths.
    _force_set(wacc, "D34", "=(D27*D16)+(D28*D21)")

    scenario_bull_beta = _to_float(_scenario_assumption_value(payload, SCENARIO_BULL, "beta"))
    scenario_bear_beta = _to_float(_scenario_assumption_value(payload, SCENARIO_BEAR, "beta"))
    base_beta = _to_float(wacc_assumptions.get("beta"))
    if base_beta is None:
        base_beta = _to_float(assumptions.get("beta"))

    _safe_set(wacc, "B34", "Base WACC")
    _safe_set(wacc, "B35", "Bull WACC")
    _safe_set(wacc, "B36", "Bear WACC")
    # Keep Base/Bull/Bear WACC rows visually consistent (fill, borders, font).
    for row in (35, 36):
        wacc[f"B{row}"]._style = copy(wacc["B34"]._style)
        wacc[f"C{row}"]._style = copy(wacc["C34"]._style)
        wacc[f"D{row}"]._style = copy(wacc["D34"]._style)
    _safe_set(wacc, "B37", "Bull Beta")
    _safe_set(wacc, "B38", "Bear Beta")
    _safe_set(wacc, "B39", "Bull Cost of Equity")
    _safe_set(wacc, "B40", "Bear Cost of Equity")
    _force_set(wacc, "C40", "")

    if scenario_bull_beta is not None:
        _force_set(wacc, "D37", max(0.10, scenario_bull_beta))
    elif base_beta is not None:
        _force_set(wacc, "D37", max(0.10, base_beta - 0.10))
    else:
        _force_set(wacc, "D37", "=MAX(0.10,D11-0.10)")

    if scenario_bear_beta is not None:
        _force_set(wacc, "D38", max(0.10, scenario_bear_beta))
    elif base_beta is not None:
        _force_set(wacc, "D38", max(0.10, base_beta + 0.10))
    else:
        _force_set(wacc, "D38", "=D11+0.10")

    _force_set(wacc, "D39", "=D9+D37*(D10)+D14+D15")
    _force_set(wacc, "D40", "=D9+D38*(D10)+D14+D15")
    _force_set(wacc, "D35", "=(D27*D39)+(D28*D21)")
    _force_set(wacc, "D36", "=(D27*D40)+(D28*D21)")

    # Harden peer beta table against partial/missing comp rows so D/E and
    # unlevered beta sections do not surface #DIV/0! in exported workbooks.
    for row in range(8, 14):
        # Debt units in the template are in billions while equity values are in millions.
        # Scale debt by 1,000 so D/E is computed on consistent units.
        _force_set(wacc, f"K{row}", f'=IFERROR(IF(J{row}>0,(I{row}*1000)/J{row},""),"")')
        _force_set(wacc, f"M{row}", f'=IFERROR(H{row}/(1+(1-L{row})*K{row}),"")')


def _harden_growth_rate_formulas(*scenario_sheets: Worksheet) -> None:
    # Guard CAGR calculations against divide-by-zero in low-data scenarios.
    formula_map = {
        "S24": "=IFERROR((I24/H24)^(1/(COLUMNS(H24:I24)-1))-1,0)",
        "T24": "=IFERROR((Q24/J24)^(1/(COLUMNS(J24:Q24)-1))-1,0)",
        "S27": "=IFERROR((I27/H27)^(1/(COLUMNS(H27:I27)-1))-1,0)",
        "T27": "=IFERROR((Q27/J27)^(1/(COLUMNS(J27:Q27)-1))-1,0)",
    }
    for sheet in scenario_sheets:
        for cell_ref, formula in formula_map.items():
            _force_set(sheet, cell_ref, formula)
        # Harden any remaining CAGR-style formulas in S/T columns that still
        # divide by historical anchors without IFERROR wrappers.
        for row in range(20, 131):
            for col in ("S", "T"):
                cell_ref = f"{col}{row}"
                value = sheet[cell_ref].value
                if not (isinstance(value, str) and value.startswith("=")):
                    continue
                upper_value = value.upper()
                if "IFERROR(" in upper_value:
                    continue
                if "COLUMNS(" not in upper_value:
                    continue
                expression = value[1:].lstrip("+")
                _force_set(sheet, cell_ref, f"=IFERROR({expression},0)")


def _normalize_dcf_waterfall_formulas(*scenario_sheets: Worksheet) -> None:
    # Keep sign conventions intuitive while preserving economics:
    # OpEx rows positive, EBIT subtracts OpEx; CapEx/NWC assumptions reference fixed inputs directly.
    for sheet in scenario_sheets:
        for col in DCF_TIMELINE_COLUMNS:
            _force_set(sheet, f"{col}48", f"={col}36+{col}39+{col}42+{col}45")
            _force_set(sheet, f"{col}49", f"=IFERROR({col}48/{col}20,0)")
            _force_set(sheet, f"{col}51", f"={col}32-{col}48")
            _force_set(sheet, f"{col}65", "=-$F$9")
            _force_set(sheet, f"{col}77", "=-$F$10")
            _force_set(sheet, f"{col}69", "=$F$13")


def _normalize_public_dcf_layout(
    outputs: Worksheet,
    dcf_base: Worksheet,
    dcf_bull: Worksheet,
    dcf_bear: Worksheet,
    payload: dict[str, Any],
    divisor: float,
) -> None:
    _link_dcf_income_statement_to_recalculated_data(dcf_base, dcf_bull, dcf_bear)
    _normalize_public_dcf_assumption_block(dcf_base, dcf_bull, dcf_bear, payload, divisor)
    _enforce_core_public_dcf_formulas(dcf_base, dcf_bull, dcf_bear)
    _enforce_outputs_bridge_formulas(outputs)


def _link_dcf_income_statement_to_recalculated_data(*scenario_sheets: Worksheet) -> None:
    # Single source of truth for DCF: annualized recalc data sheet.
    for sheet in scenario_sheets:
        for idx, dcf_col in enumerate(DCF_TIMELINE_COLUMNS):
            recalc_col = RECALC_COLUMNS[idx]
            _force_set(sheet, f"{dcf_col}20", f"='{SHEET_DATA_RECALCULATED}'!{recalc_col}12")
            _force_set(sheet, f"{dcf_col}24", f"='{SHEET_DATA_RECALCULATED}'!{recalc_col}16")
            _force_set(sheet, f"{dcf_col}27", f"='{SHEET_DATA_RECALCULATED}'!{recalc_col}17")
            _force_set(sheet, f"{dcf_col}36", f"='{SHEET_DATA_RECALCULATED}'!{recalc_col}24")
            _force_set(sheet, f"{dcf_col}39", f"='{SHEET_DATA_RECALCULATED}'!{recalc_col}25")
            _force_set(sheet, f"{dcf_col}42", f"='{SHEET_DATA_RECALCULATED}'!{recalc_col}26")
            _force_set(sheet, f"{dcf_col}45", f"='{SHEET_DATA_RECALCULATED}'!{recalc_col}27")


def _normalize_public_dcf_assumption_block(
    dcf_base: Worksheet,
    dcf_bull: Worksheet,
    dcf_bear: Worksheet,
    payload: dict[str, Any],
    divisor: float,
) -> None:
    market = payload.get("market", {})
    cash = _to_float(market.get("cash")) or 0.0
    debt = _to_float(market.get("debt")) or 0.0
    non_operating_assets = _to_float(market.get("nonOperatingAssets")) or 0.0

    for sheet, wacc_cell in (
        (dcf_base, "D34"),
        (dcf_bull, "D35"),
        (dcf_bear, "D36"),
    ):
        _safe_set(sheet, "B8", "Valuation Inputs")
        _safe_set(sheet, "B9", "Enterprise Value")
        _safe_set(sheet, "B10", "EV / EBITDA (LTM)")
        _safe_set(sheet, "B11", "Current Equity Value")
        _safe_set(sheet, "B12", "Implied Equity Value")
        _safe_set(sheet, "E14", "Revenue Growth Rate")
        _safe_set(sheet, "B15", "Terminal Assumptions")
        _safe_set(sheet, "B16", "Exit EBITDA Multiple")
        _safe_set(sheet, "B17", "Cash and Cash Equivalents")
        _safe_set(sheet, "B18", "Income Statement")
        _safe_set_or_clear(sheet, "B19", None)
        _safe_set_or_clear(sheet, "B13", None)
        _safe_set_or_clear(sheet, "C13", None)
        _safe_set_or_clear(sheet, "B14", None)
        _safe_set_or_clear(sheet, "C14", None)
        sheet["F14"]._style = copy(sheet["F11"]._style)
        _safe_set_or_clear(sheet, "F14", _to_float(sheet["F14"].value) or _to_float(sheet["C14"].value) or _to_float(sheet["C13"].value))
        sheet["F14"].number_format = "0.0%"
        _safe_set_or_clear(sheet, "E15", None)
        _safe_set_or_clear(sheet, "E16", None)
        _safe_set_or_clear(sheet, "E19", None)
        _force_set(sheet, "C17", _scale(cash, divisor))
        _force_set(sheet, "C10", "=IFERROR(C9/L54,0)")
        _force_set(sheet, "C12", f"=C9-{DCF_HELPER_DEBT_CELL}+{DCF_HELPER_CASH_CELL}+{DCF_HELPER_NON_OP_CELL}")
        _force_set(sheet, "F12", f"=WACC!{wacc_cell}")
        _safe_set_or_clear(sheet, "F16", None)
        _safe_set_or_clear(sheet, "F17", None)
        _safe_set_or_clear(sheet, "F18", None)
        _safe_set_or_clear(sheet, "F19", None)
        _force_set(sheet, DCF_HELPER_CASH_CELL, "=C17")
        _force_set(sheet, DCF_HELPER_DEBT_CELL, _scale(debt, divisor))
        _force_set(sheet, DCF_HELPER_NON_OP_CELL, _scale(non_operating_assets, divisor))
        for assumption_ref in ("F9", "F10", "F11", "F12", "F13", "F14", "C16"):
            sheet[assumption_ref].comment = None


def _enforce_core_public_dcf_formulas(*scenario_sheets: Worksheet) -> None:
    projection_columns = DCF_TIMELINE_COLUMNS[2:]  # J..Q
    formula_projection_columns = DCF_TIMELINE_COLUMNS[5:]  # M..Q

    for sheet in scenario_sheets:
        for col in DCF_TIMELINE_COLUMNS:
            _force_set(sheet, f"{col}30", f"={col}24+{col}27")
            _force_set(sheet, f"{col}32", f"={col}20-{col}30")
            _force_set(sheet, f"{col}33", f"=IFERROR({col}32/{col}20,0)")
            _force_set(sheet, f"{col}54", f"={col}51+{col}68")
            _force_set(sheet, f"{col}55", f"=IFERROR({col}54/{col}20,0)")
            _force_set(sheet, f"{col}57", f"=-{col}51*{col}58")
            _force_set(sheet, f"{col}58", "=$F$11")
            _force_set(sheet, f"{col}60", f"={col}51+{col}57")
            _force_set(sheet, f"{col}68", f"={col}69*{col}20")
            _force_set(sheet, f"{col}70", f"=IFERROR(-{col}68/{col}65,0)")
            _force_set(sheet, f"{col}74", f"={col}60")
            _force_set(sheet, f"{col}75", f"={col}68")
            _force_set(sheet, f"{col}76", f"={col}65")
            _force_set(sheet, f"{col}78", f"=SUM({col}74:{col}77)")
            _force_set(sheet, f"{col}79", f"=IFERROR({col}78/{col}20,0)")
            _force_set(sheet, f"{col}85", "=$F$12")
            if col not in formula_projection_columns:
                _force_set(sheet, f"{col}25", f"=IFERROR({col}24/{col}20,0)")
                _force_set(sheet, f"{col}37", f"=IFERROR({col}36/{col}20,0)")
                _force_set(sheet, f"{col}40", f"=IFERROR({col}39/{col}20,0)")
                _force_set(sheet, f"{col}43", f"=IFERROR({col}42/{col}20,0)")
                _force_set(sheet, f"{col}46", f"=IFERROR({col}45/{col}20,0)")

        for idx, col in enumerate(formula_projection_columns):
            prev_col = DCF_TIMELINE_COLUMNS[5 + idx - 1]
            _force_set(sheet, f"{col}25", f"={prev_col}25")
            _force_set(sheet, f"{col}37", f"={prev_col}37")
            _force_set(sheet, f"{col}40", f"={prev_col}40")
            _force_set(sheet, f"{col}43", f"={prev_col}43")
            _force_set(sheet, f"{col}46", f"={prev_col}46")
            _force_set(sheet, f"{col}24", f"={col}20*{col}25")
            _force_set(sheet, f"{col}36", f"={col}20*{col}37")
            _force_set(sheet, f"{col}39", f"={col}20*{col}40")
            _force_set(sheet, f"{col}42", f"={col}20*$F$13")
            _force_set(sheet, f"{col}45", f"={col}20*{col}46")
            _force_set(sheet, f"{col}65", "=-$F$9")
            _force_set(sheet, f"{col}77", "=-$F$10")

        _force_set(sheet, "E81", "=I9")
        for idx, col in enumerate(projection_columns):
            prev_col = projection_columns[idx - 1] if idx > 0 else None
            _force_set(sheet, f"{col}82", "=I11" if col == "J" else f"=EOMONTH({prev_col}82,12)")
            _force_set(sheet, f"{col}83", f"=({col}82-E81)/365" if col == "J" else f"={prev_col}83+1")
            _force_set(sheet, f"{col}84", f"={col}83/2" if col == "J" else f"={col}83-0.5")
            _force_set(sheet, f"{col}87", f"={col}78/(1+{col}85)^{col}84")

        _force_set(sheet, "Q92", "=Q54*$C$16")
        _force_set(sheet, "Q93", "=Q92/(1+Q85)^Q83")
        _force_set(sheet, "Q94", "=Q93+SUM(J87:Q87)")
        _force_set(sheet, "C9", "=Q94")
        _force_set(sheet, "Q95", f"=-{DCF_HELPER_DEBT_CELL_ABS}")
        _force_set(sheet, "Q96", f"={DCF_HELPER_NON_OP_CELL_ABS}")
        _force_set(sheet, "Q97", f"={DCF_HELPER_CASH_CELL_ABS}")
        _force_set(sheet, "Q98", "=SUM(Q94:Q97)")
        _force_set(sheet, "Q99", "=C11")
        _force_set(sheet, "Q100", "=IFERROR(Q98/Q99 - 1,0)")
        _force_set(sheet, "Q104", "=Q78*(1+Q103)")
        _force_set(sheet, "Q105", "=IFERROR(IF(Q85>Q103,Q104/(Q85-Q103),0),0)")
        _force_set(sheet, "Q106", "=Q105/(1+Q85)^Q83")
        _force_set(sheet, "Q107", "=Q106+SUM(J87:Q87)")
        _force_set(sheet, "Q108", f"=-{DCF_HELPER_DEBT_CELL_ABS}")
        _force_set(sheet, "Q109", f"={DCF_HELPER_NON_OP_CELL_ABS}")
        _force_set(sheet, "Q110", f"={DCF_HELPER_CASH_CELL_ABS}")
        _force_set(sheet, "Q111", "=SUM(Q107:Q110)")
        _force_set(sheet, "Q112", "=C11")
        _force_set(sheet, "Q113", "=IFERROR(Q111/Q112 - 1,0)")

        first_forecast_idx = next(
            (idx for idx, col in enumerate(DCF_TIMELINE_COLUMNS) if isinstance(sheet[f"{col}18"].value, str) and str(sheet[f"{col}18"].value).endswith("E")),
            None,
        )
        if first_forecast_idx is not None:
            anchor_col = DCF_TIMELINE_COLUMNS[first_forecast_idx]
            sheet[f"{anchor_col}20"].comment = None
            for idx in range(first_forecast_idx + 1, len(DCF_TIMELINE_COLUMNS)):
                col = DCF_TIMELINE_COLUMNS[idx]
                prev_col = DCF_TIMELINE_COLUMNS[idx - 1]
                _force_set(sheet, f"{col}20", f"=IFERROR({prev_col}20*(1+$F$14),{prev_col}20)")


def _scenario_choose_formula(base_cell: str, bull_cell: str | None = None, bear_cell: str | None = None) -> str:
    bull_ref = bull_cell or base_cell
    bear_ref = bear_cell or base_cell
    return (
        "=CHOOSE(Cover!$C$12,"
        f"'{SHEET_DCF_BASE}'!{base_cell},"
        f"'{SHEET_DCF_BULL}'!{bull_ref},"
        f"'{SHEET_DCF_BEAR}'!{bear_ref})"
    )


def _enforce_outputs_bridge_formulas(outputs: Worksheet) -> None:
    _force_set(outputs, "E18", _scenario_choose_formula("I9"))
    _force_set(outputs, "J19", _scenario_choose_formula("I11"))
    _force_set(outputs, "J20", "=(J19-E18)/365")
    _force_set(outputs, "J21", "=J20/2")
    _force_set(outputs, "D27", _scenario_choose_formula("$F$12"))
    _force_set(outputs, "H27", "=D27")
    _force_set(outputs, "D28", _scenario_choose_formula("$C$16"))
    _force_set(outputs, "H28", _scenario_choose_formula("$Q$103"))

    for col in ("H", "I", "J", "K", "L", "M", "N", "O", "P", "Q"):
        _force_set(outputs, f"{col}8", _scenario_choose_formula(f"{col}51"))
        _force_set(outputs, f"{col}9", _scenario_choose_formula(f"{col}68"))
        _force_set(outputs, f"{col}10", _scenario_choose_formula(f"{col}54"))
        _force_set(outputs, f"{col}12", _scenario_choose_formula(f"{col}60"))
        _force_set(outputs, f"{col}13", f"={col}9")
        _force_set(outputs, f"{col}14", _scenario_choose_formula(f"{col}65"))
        _force_set(outputs, f"{col}15", f"=-{_scenario_choose_formula('$F$10')[1:]}")
        _force_set(outputs, f"{col}22", _scenario_choose_formula("$F$12"))

    _safe_set(outputs, "B37", "(+) Non-Operating Assets")
    _safe_set(outputs, "F37", "(+) Non-Operating Assets")
    _force_set(outputs, "D33", "=D32/(1+D27)^Q20")
    _force_set(outputs, "H33", "=H32/(1+H27)^Q20")
    _force_set(outputs, "D36", f"=-{_scenario_choose_formula(DCF_HELPER_DEBT_CELL_ABS)[1:]}")
    _force_set(outputs, "H36", "=D36")
    _force_set(outputs, "D37", _scenario_choose_formula(DCF_HELPER_NON_OP_CELL_ABS))
    _force_set(outputs, "H37", "=D37")
    _force_set(outputs, "D38", _scenario_choose_formula(DCF_HELPER_CASH_CELL_ABS))
    _force_set(outputs, "H38", "=D38")
    _force_set(outputs, "D41", _scenario_choose_formula("$C$11"))
    _force_set(outputs, "H41", "=D41")


def _sync_scenario_formula_backbone(dcf_base: Worksheet, *scenario_sheets: Worksheet) -> None:
    # Keep formula topology identical across Base/Bull/Bear to avoid scenario drift.
    scenario_specific_formula_cells = {"F12"}
    for row in dcf_base.iter_rows(min_row=1, max_row=dcf_base.max_row, min_col=1, max_col=dcf_base.max_column):
        for base_cell in row:
            formula = base_cell.value
            if not (isinstance(formula, str) and formula.startswith("=")):
                continue
            if base_cell.coordinate in scenario_specific_formula_cells:
                continue
            for scenario in scenario_sheets:
                scenario_cell = scenario[base_cell.coordinate]
                if isinstance(scenario_cell, MergedCell):
                    continue
                scenario_cell.value = formula


def _apply_capex_schedule_to_dcf(
    dcf_base: Worksheet,
    dcf_bull: Worksheet,
    dcf_bear: Worksheet,
    payload: dict[str, Any],
    timeline_years: list[int],
    divisor: float,
) -> None:
    # Projection-period CapEx rows are formula-driven from assumptions.
    # Avoid stamping hardcoded year-by-year values that can drift.
    return


def _scenario_snapshot(payload: dict[str, Any], scenario_name: str) -> dict[str, Any] | None:
    scenarios = payload.get("scenarios")
    if not isinstance(scenarios, dict):
        return None
    snapshot = scenarios.get(scenario_name)
    return snapshot if isinstance(snapshot, dict) else None


def _forecast_map(snapshot: dict[str, Any]) -> dict[int, dict[str, Any]]:
    forecasts = snapshot.get("forecasts")
    if not isinstance(forecasts, list):
        return {}

    out: dict[int, dict[str, Any]] = {}
    for forecast in forecasts:
        if not isinstance(forecast, dict):
            continue
        year = _to_float(forecast.get("year"))
        if year is None:
            continue
        out[int(year)] = forecast
    return out


def _scenario_assumptions(snapshot: dict[str, Any]) -> dict[str, Any]:
    assumptions = snapshot.get("assumptions")
    return assumptions if isinstance(assumptions, dict) else {}


def _scenario_summary(snapshot: dict[str, Any]) -> dict[str, Any]:
    summary = snapshot.get("summary")
    return summary if isinstance(summary, dict) else {}


def _scenario_first_projection_forecast(
    forecast_by_year: dict[int, dict[str, Any]],
    timeline_years: list[int],
) -> dict[str, Any] | None:
    projection_years = timeline_years[2:] if len(timeline_years) >= 3 else []
    for year in projection_years:
        forecast = forecast_by_year.get(year)
        if forecast is not None:
            return forecast
    if not forecast_by_year:
        return None
    first_year = sorted(forecast_by_year.keys())[0]
    return forecast_by_year[first_year]


def _map_scenario_forecasts_to_sheet(
    sheet: Worksheet,
    forecast_by_year: dict[int, dict[str, Any]],
    timeline_years: list[int],
    divisor: float,
) -> None:
    projection_columns = DCF_TIMELINE_COLUMNS[2:]  # J..Q
    projection_years = timeline_years[2:] if len(timeline_years) >= 3 else []
    first_projection_year = next((year for year in projection_years if year in forecast_by_year), None)

    for idx, col in enumerate(projection_columns):
        if idx >= len(projection_years):
            break
        projection_year = projection_years[idx]
        forecast = forecast_by_year.get(projection_year)
        if not isinstance(forecast, dict):
            continue

        revenue = _to_float(forecast.get("revenue"))
        # Keep first projected revenue as an anchor input; later projected years
        # should compound from the explicit growth-rate assumption formula.
        if revenue is not None and projection_year == first_projection_year:
            _force_set(sheet, f"{col}20", _scale(revenue, divisor))


def _apply_scenario_snapshot_to_sheet(
    sheet: Worksheet,
    snapshot: dict[str, Any],
    timeline_years: list[int],
    divisor: float,
) -> None:
    assumptions = _scenario_assumptions(snapshot)
    forecast_by_year = _forecast_map(snapshot)

    tax_rate = _to_float(assumptions.get("taxRate"))
    wacc_rate = _sanitize_wacc_rate(assumptions.get("waccRate"))
    da_pct_revenue = _to_float(assumptions.get("daPctRevenue"))
    revenue_growth_rate = _to_float(assumptions.get("revenueGrowthRate"))
    if revenue_growth_rate is None:
        revenue_growth_rate = _to_float(assumptions.get("revenueGrowth"))
    if revenue_growth_rate is None:
        revenue_growth_rate = _infer_revenue_growth_rate(snapshot.get("forecasts") if isinstance(snapshot.get("forecasts"), list) else [])
    terminal_growth = _sanitize_terminal_growth_rate(
        assumptions.get("terminalGrowthRate"),
        reference_wacc=wacc_rate,
    )
    exit_multiple = _to_float(assumptions.get("terminalExitMultiple"))

    if tax_rate is not None:
        _force_set(sheet, "F11", tax_rate)
    if da_pct_revenue is not None:
        _force_set(sheet, "F13", da_pct_revenue)
    if revenue_growth_rate is not None:
        _force_set(sheet, "F14", revenue_growth_rate)
        sheet["F14"].number_format = "0.0%"
    if terminal_growth is not None:
        _force_set(sheet, "Q103", terminal_growth)
    if exit_multiple is not None:
        _force_set(sheet, "C16", exit_multiple)

    first_projection = _scenario_first_projection_forecast(forecast_by_year, timeline_years)
    if isinstance(first_projection, dict):
        capex = _to_float(first_projection.get("capex"))
        nwc_change = _first_float(first_projection, "nwcChange", "nwc_change")
        if capex is not None:
            _force_set(sheet, "F9", _scale(abs(capex), divisor))
        if nwc_change is not None:
            _force_set(sheet, "F10", _scale(abs(nwc_change), divisor))

    _map_scenario_forecasts_to_sheet(sheet, forecast_by_year, timeline_years, divisor)


def _apply_scenario_snapshots_to_dcf(
    dcf_base: Worksheet,
    dcf_bull: Worksheet,
    dcf_bear: Worksheet,
    payload: dict[str, Any],
    timeline_years: list[int],
    divisor: float,
) -> None:
    base_snapshot = _scenario_snapshot(payload, SCENARIO_BASE)
    bull_snapshot = _scenario_snapshot(payload, SCENARIO_BULL)
    bear_snapshot = _scenario_snapshot(payload, SCENARIO_BEAR)

    if base_snapshot is None and bull_snapshot is None and bear_snapshot is None:
        return

    if base_snapshot is not None:
        _apply_scenario_snapshot_to_sheet(dcf_base, base_snapshot, timeline_years, divisor)
    if bull_snapshot is not None:
        _apply_scenario_snapshot_to_sheet(dcf_bull, bull_snapshot, timeline_years, divisor)
    if bear_snapshot is not None:
        _apply_scenario_snapshot_to_sheet(dcf_bear, bear_snapshot, timeline_years, divisor)


def _finalize_assumption_block_cleanup(*scenario_sheets: Worksheet) -> None:
    # Ensure assumption area is clean and free of duplicate helper numbers.
    for sheet in scenario_sheets:
        _safe_set(sheet, "B18", "Income Statement")
        _safe_set_or_clear(sheet, "B19", None)
        _safe_set_or_clear(sheet, "E15", None)
        _safe_set_or_clear(sheet, "E16", None)
        _safe_set_or_clear(sheet, "F16", None)
        _safe_set_or_clear(sheet, "F17", None)
        _safe_set_or_clear(sheet, "F18", None)
        _safe_set_or_clear(sheet, "F19", None)
        _safe_set_or_clear(sheet, "E19", None)
        for address in ("C18", "C19", "D18", "D19", "E18", "G18", "G19"):
            _safe_set_or_clear(sheet, address, None)
        # Remove lingering note indicators (red triangles) from template/input mapping.
        for address in (
            "C13",
            "C16",
            "C17",
            "E15",
            "E16",
            "F9",
            "F10",
            "F11",
            "F12",
            "F13",
            "F14",
            "F15",
            "F16",
            "F17",
            "F18",
            "F19",
        ):
            sheet[address].comment = None


def _finalize_timeline_headers(
    outputs: Worksheet,
    dcf_base: Worksheet,
    dcf_bull: Worksheet,
    dcf_bear: Worksheet,
    timeline_years: list[int],
    historical_years: set[int],
) -> None:
    # Final guardrail: force all timeline headers to explicit FY labels as text.
    for idx, col in enumerate(DCF_TIMELINE_COLUMNS):
        year = timeline_years[idx]
        label = f"FY{year}{'A' if year in historical_years else 'E'}"
        _force_set(outputs, f"{col}6", label)
        outputs[f"{col}6"].number_format = "@"
        for sheet in (dcf_base, dcf_bull, dcf_bear):
            _force_set(sheet, f"{col}18", label)
            _force_set(sheet, f"{col}63", label)
            _force_set(sheet, f"{col}72", label)
            _force_set(sheet, f"{col}89", label)
            sheet[f"{col}89"].number_format = "@"


def _reset_dcf_sheet_view_to_top(*sheets: Worksheet) -> None:
    # Ensure each scenario tab opens at the top of the sheet instead of
    # preserving a template viewport near the bottom.
    for sheet in sheets:
        sheet.sheet_view.topLeftCell = "A1"
        if sheet.sheet_view.selection:
            sheet.sheet_view.selection[0].activeCell = "A1"
            sheet.sheet_view.selection[0].sqref = "A1"


def _numeric_list(raw: Any) -> list[float]:
    if not isinstance(raw, list):
        return []
    out: list[float] = []
    for item in raw:
        parsed = _to_float(item)
        if parsed is not None:
            out.append(parsed)
    return out


def _matrix_values(raw: Any) -> list[list[float]] | None:
    if not isinstance(raw, list) or len(raw) != 5:
        return None
    out: list[list[float]] = []
    for row in raw:
        numeric_row = _numeric_list(row)
        if len(numeric_row) != 5:
            return None
        out.append(numeric_row)
    return out


def _axis_from_bounds(
    values: list[float],
    *,
    center: float,
    step: float,
    min_value: float,
    max_value: float,
) -> list[float]:
    if len(values) >= 2:
        low = max(min_value, min(values))
        high = min(max_value, max(values))
        if high <= low:
            high = min(max_value, low + step * 4)
        if high > low:
            return [round(low + (high - low) * idx / 4, 4) for idx in range(5)]

    seed = center if center == center else (min_value + max_value) / 2
    axis = [seed + (idx - 2) * step for idx in range(5)]
    clamped = [max(min_value, min(max_value, value)) for value in axis]
    for idx in range(1, len(clamped)):
        if clamped[idx] <= clamped[idx - 1]:
            clamped[idx] = min(max_value, clamped[idx - 1] + max(step / 2, 0.0005))
    return [round(value, 4) for value in clamped]


def _format_percent_axis_label(value: float) -> str:
    return f"{value * 100:.1f}%"


def _uniform_axis_step(values: list[float], *, tolerance: float = 1e-6) -> float | None:
    if len(values) != 5:
        return None
    deltas = [values[idx + 1] - values[idx] for idx in range(len(values) - 1)]
    if any(delta <= 0 for delta in deltas):
        return None
    first = deltas[0]
    if all(abs(delta - first) <= tolerance for delta in deltas[1:]):
        return first
    return None


def _set_percent_axis_row(
    sheet: Worksheet,
    *,
    cells: tuple[str, str, str, str, str],
    values: list[float],
) -> None:
    step = _uniform_axis_step(values)
    if step is not None:
        center = cells[2]
        _force_set(sheet, center, values[2])
        _force_set(sheet, cells[1], f"={center}-{step:.4f}")
        _force_set(sheet, cells[0], f"={cells[1]}-{step:.4f}")
        _force_set(sheet, cells[3], f"={center}+{step:.4f}")
        _force_set(sheet, cells[4], f"={cells[3]}+{step:.4f}")
    else:
        for cell_ref, value in zip(cells, values, strict=False):
            _force_set(sheet, cell_ref, value)

    for cell_ref in cells:
        sheet[cell_ref].number_format = "0.0%"


def _set_percent_axis_column(
    sheet: Worksheet,
    *,
    cells: tuple[str, str, str, str, str],
    values: list[float],
) -> None:
    step = _uniform_axis_step(values)
    if step is not None:
        center = cells[2]
        _force_set(sheet, center, values[2])
        _force_set(sheet, cells[1], f"={center}-{step:.4f}")
        _force_set(sheet, cells[0], f"={cells[1]}-{step:.4f}")
        _force_set(sheet, cells[3], f"={center}+{step:.4f}")
        _force_set(sheet, cells[4], f"={cells[3]}+{step:.4f}")
    else:
        for cell_ref, value in zip(cells, values, strict=False):
            _force_set(sheet, cell_ref, value)

    for cell_ref in cells:
        sheet[cell_ref].number_format = "0.0%"


def _fallback_wacc_terminal_matrix(
    *,
    base_ev: float,
    base_wacc: float,
    base_growth: float,
    wacc_axis: list[float],
    growth_axis: list[float],
    tv_weight: float,
) -> list[list[float]]:
    explicit_component = base_ev * (1.0 - tv_weight)
    terminal_component = base_ev * tv_weight
    base_spread = max(0.005, base_wacc - base_growth)

    matrix: list[list[float]] = []
    for growth in growth_axis:
        row: list[float] = []
        for wacc in wacc_axis:
            spread = max(0.005, wacc - growth)
            spread_factor = base_spread / spread
            explicit_factor = max(0.5, min(1.5, 1.0 - 1.5 * (wacc - base_wacc)))
            value = (explicit_component * explicit_factor) + (terminal_component * spread_factor)
            row.append(max(1.0, value))
        matrix.append(row)
    matrix[2][2] = base_ev
    return matrix


def _fallback_revenue_ebit_matrix(
    *,
    base_ev: float,
    base_revenue_growth: float,
    base_ebit_margin: float,
    revenue_growth_axis: list[float],
    ebit_margin_axis: list[float],
) -> list[list[float]]:
    matrix: list[list[float]] = []
    for margin in ebit_margin_axis:
        row: list[float] = []
        for growth in revenue_growth_axis:
            growth_factor = max(0.3, 1.0 + 4.0 * (growth - base_revenue_growth))
            margin_factor = max(0.3, 1.0 + 6.0 * (margin - base_ebit_margin))
            row.append(max(1.0, base_ev * growth_factor * margin_factor))
        matrix.append(row)

    matrix[2][2] = base_ev
    for row_idx in range(5):
        for col_idx in range(5):
            if col_idx > 0:
                matrix[row_idx][col_idx] = max(matrix[row_idx][col_idx], matrix[row_idx][col_idx - 1])
            if row_idx > 0:
                matrix[row_idx][col_idx] = max(matrix[row_idx][col_idx], matrix[row_idx - 1][col_idx])
    return matrix


def _map_sensitivity_blocks(
    dcf_base: Worksheet,
    dcf_bull: Worksheet,
    dcf_bear: Worksheet,
    payload: dict[str, Any],
    divisor: float,
) -> None:
    assumptions = payload.get("assumptions", {})
    assumptions = assumptions if isinstance(assumptions, dict) else {}
    terminal = assumptions.get("terminal", {})
    terminal = terminal if isinstance(terminal, dict) else {}
    sensitivities = payload.get("sensitivities", {})
    sensitivities = sensitivities if isinstance(sensitivities, dict) else {}

    key_metrics = (payload.get("uiMeta") or {}).get("keyMetrics") or {}
    market = payload.get("market", {})
    shares = _to_float(market.get("sharesDiluted")) or 0.0
    price = _to_float(market.get("currentPrice")) or 0.0
    market_cap = shares * price if shares > 0 and price > 0 else None
    net_debt = _to_float(market.get("netDebt"))
    if net_debt is None:
        debt = _to_float(market.get("debt")) or 0.0
        cash = _to_float(market.get("cash")) or 0.0
        net_debt = debt - cash

    base_ev_raw = _to_float(key_metrics.get("enterpriseValue"))
    if base_ev_raw is None and market_cap is not None:
        base_ev_raw = market_cap + (net_debt or 0.0)
    if base_ev_raw is None:
        base_ev_raw = 1_000_000.0

    pv_terminal_raw = _to_float(key_metrics.get("pvTerminalValue")) or 0.0
    tv_weight = pv_terminal_raw / base_ev_raw if base_ev_raw > 0 else 0.7
    tv_weight = max(0.4, min(0.9, tv_weight))

    base_wacc = _sanitize_wacc_rate(assumptions.get("waccRate")) or 0.10
    base_growth = _sanitize_terminal_growth_rate(terminal.get("g"), reference_wacc=base_wacc) or 0.025
    base_revenue_growth = _to_float(assumptions.get("revenueGrowth")) or 0.03
    base_ebit_margin = _to_float(assumptions.get("ebitMargin")) or _to_float(assumptions.get("ebitMarginTarget")) or 0.15

    wacc_axis = _axis_from_bounds(
        _numeric_list(sensitivities.get("waccAxis")) or _numeric_list(sensitivities.get("waccGrid")),
        center=base_wacc,
        step=0.01,
        min_value=0.01,
        max_value=0.30,
    )
    growth_axis = _axis_from_bounds(
        _numeric_list(sensitivities.get("terminalGrowthAxis")) or _numeric_list(sensitivities.get("gGrid")),
        center=base_growth,
        step=0.005,
        min_value=0.0,
        max_value=max(0.001, min(0.08, min(wacc_axis) - 0.001)),
    )
    revenue_growth_axis = _axis_from_bounds(
        _numeric_list(sensitivities.get("revenueGrowthAxis")),
        center=base_revenue_growth,
        step=0.01,
        min_value=-0.10,
        max_value=0.30,
    )
    ebit_margin_axis = _axis_from_bounds(
        _numeric_list(sensitivities.get("ebitMarginAxis")),
        center=base_ebit_margin,
        step=0.01,
        min_value=0.01,
        max_value=0.60,
    )

    wacc_terminal_matrix = _matrix_values(sensitivities.get("waccTerminalEvMatrix"))
    if wacc_terminal_matrix is None:
        wacc_terminal_matrix = _fallback_wacc_terminal_matrix(
            base_ev=base_ev_raw,
            base_wacc=base_wacc,
            base_growth=base_growth,
            wacc_axis=wacc_axis,
            growth_axis=growth_axis,
            tv_weight=tv_weight,
        )

    revenue_ebit_matrix = _matrix_values(sensitivities.get("revenueEbitEvMatrix"))
    if revenue_ebit_matrix is None:
        revenue_ebit_matrix = _fallback_revenue_ebit_matrix(
            base_ev=base_ev_raw,
            base_revenue_growth=base_revenue_growth,
            base_ebit_margin=base_ebit_margin,
            revenue_growth_axis=revenue_growth_axis,
            ebit_margin_axis=ebit_margin_axis,
        )

    for sheet in (dcf_base, dcf_bull, dcf_bear):
        # Add visual space between the two sensitivity tables.
        current_i_width = sheet.column_dimensions["I"].width
        if current_i_width is None or current_i_width < 14:
            sheet.column_dimensions["I"].width = 14

        _safe_set(sheet, "C117", "Enterprise Value - WACC x Terminal Growth")
        _safe_set(sheet, "D118", "WACC")
        _force_set(sheet, "C119", f"='{SHEET_OUTPUTS}'!$D$35")
        _safe_set(sheet, "B121", "Terminal Growth")
        _safe_set(sheet, "B122", "Rate")
        _set_percent_axis_row(sheet, cells=("D119", "E119", "F119", "G119", "H119"), values=wacc_axis)
        _set_percent_axis_column(sheet, cells=("C120", "C121", "C122", "C123", "C124"), values=growth_axis)

        for row_idx in range(5):
            for col_idx in range(5):
                scaled = _scale(wacc_terminal_matrix[row_idx][col_idx], divisor)
                _force_set(sheet, f"{chr(ord('D') + col_idx)}{120 + row_idx}", scaled)

        _force_set(sheet, "C126", "=MIN(D120:H124)")
        _force_set(sheet, "C127", "=PERCENTILE(D120:H124,0.25)")
        _force_set(sheet, "C128", "=MEDIAN(D120:H124)")
        _force_set(sheet, "C129", "=PERCENTILE(D120:H124,0.75)")
        _force_set(sheet, "C130", "=MAX(D120:H124)")
        _safe_set(sheet, "B126", "Min")
        _safe_set(sheet, "B127", "Q1")
        _safe_set(sheet, "B128", "Median")
        _safe_set(sheet, "B129", "Q3")
        _safe_set(sheet, "B130", "Max")

        # Ensure every right-side table cell is materialized in sheet XML.
        for row in range(117, 131):
            for col in range(ord("I"), ord("O") + 1):
                ref = f"{chr(col)}{row}"
                if sheet[ref].value is None:
                    _force_set(sheet, ref, "")

        _safe_set(sheet, "I117", "Enterprise Value - Revenue Growth x EBIT Margin")
        _safe_set(sheet, "J118", "Revenue Growth")
        _safe_set(sheet, "I121", "EBIT Margin")
        _safe_set(sheet, "I122", "Rate")
        _set_percent_axis_row(sheet, cells=("J119", "K119", "L119", "M119", "N119"), values=revenue_growth_axis)
        _set_percent_axis_column(sheet, cells=("I120", "I121", "I122", "I123", "I124"), values=ebit_margin_axis)

        for row_idx in range(5):
            for col_idx in range(5):
                scaled = _scale(revenue_ebit_matrix[row_idx][col_idx], divisor)
                numeric = scaled if scaled is not None else 0.0
                _force_set(sheet, f"{chr(ord('J') + col_idx)}{120 + row_idx}", numeric)

        _safe_set(sheet, "I126", "Min")
        _safe_set(sheet, "I127", "Q1")
        _safe_set(sheet, "I128", "Median")
        _safe_set(sheet, "I129", "Q3")
        _safe_set(sheet, "I130", "Max")
        _force_set(sheet, "J126", "=MIN(J120:N124)")
        _force_set(sheet, "J127", "=PERCENTILE(J120:N124,0.25)")
        _force_set(sheet, "J128", "=MEDIAN(J120:N124)")
        _force_set(sheet, "J129", "=PERCENTILE(J120:N124,0.75)")
        _force_set(sheet, "J130", "=MAX(J120:N124)")

        # Highlight low/median/high valuation outcomes with a standard red-yellow-green gradient.
        for target_range in ("D120:H124", "J120:N124"):
            sheet.conditional_formatting.add(
                target_range,
                ColorScaleRule(
                    start_type="min",
                    start_color="F8696B",
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="FFEB84",
                    end_type="max",
                    end_color="63BE7B",
                ),
            )

    _clear_sensitivity_blocks(dcf_bull, dcf_bear)


def _clear_sensitivity_blocks(*scenario_sheets: Worksheet) -> None:
    for sheet in scenario_sheets:
        for row in range(117, 131):
            for col in "BCDEFGHIJKLMNO":
                _force_set(sheet, f"{col}{row}", None)


def _harden_comps_ratio_formulas(comps_ws: Worksheet) -> None:
    # Avoid propagating #DIV/0! when one or more peer rows are missing inputs.
    for row in range(8, 14):
        _force_set(comps_ws, f"H{row}", f'=IFERROR(P{row}/R{row},"")')
        _force_set(comps_ws, f"I{row}", f'=IFERROR($G{row}/P{row},"")')
        _force_set(comps_ws, f"J{row}", f'=IFERROR($G{row}/Q{row},"")')
        _force_set(comps_ws, f"K{row}", f'=IFERROR($G{row}/R{row},"")')
        _force_set(comps_ws, f"L{row}", f'=IFERROR($G{row}/S{row},"")')


def _harden_wacc_peer_aggregate_formulas(wacc: Worksheet) -> None:
    # Guard peer aggregates to avoid #DIV/0! when comparable set is partially empty.
    _force_set(wacc, "H16", "=IFERROR(AVERAGE(H8:H13),1)")
    _force_set(wacc, "I16", "=IFERROR(AVERAGE(I8:I13),0)")
    _force_set(wacc, "J16", "=IFERROR(AVERAGE(J8:J13),0)")
    _force_set(wacc, "K16", "=IFERROR(AVERAGE(K8:K13),0)")
    _force_set(wacc, "L16", "=IFERROR(AVERAGE(L8:L13),0.25)")
    _force_set(wacc, "M16", "=IFERROR(AVERAGE(M8:M13),1)")
    _force_set(wacc, "N16", "=IFERROR(AVERAGE(N8:N13),1)")
    _force_set(wacc, "H17", "=IFERROR(MEDIAN(H8:H13),1)")
    _force_set(wacc, "I17", "=IFERROR(MEDIAN(I8:I13),0)")
    _force_set(wacc, "J17", "=IFERROR(MEDIAN(J8:J13),0)")
    _force_set(wacc, "K17", "=IFERROR(MEDIAN(K8:K13),0)")
    _force_set(wacc, "L17", "=IFERROR(MEDIAN(L8:L13),0.25)")
    _force_set(wacc, "M17", "=IFERROR(MEDIAN(M8:M13),1)")
    _force_set(wacc, "N17", "=IFERROR(MEDIAN(N8:N13),1)")


def _normalize_comp_name(comp: dict[str, Any]) -> str | None:
    company = comp.get("company")
    if not (isinstance(company, str) and company.strip()):
        company = comp.get("name") or comp.get("companyName")
    ticker = comp.get("ticker") or comp.get("symbol")
    if isinstance(company, str) and company.strip() and isinstance(ticker, str) and ticker.strip():
        return f"{company} ({ticker})"
    if isinstance(company, str) and company.strip():
        return company
    if isinstance(ticker, str) and ticker.strip():
        return ticker
    return None


def _normalize_comp_shares(raw: Any) -> float | None:
    shares = _to_float(raw)
    if shares is None:
        return None
    return shares / 1_000_000.0 if abs(shares) >= 1_000_000 else shares


def _first_float(comp: dict[str, Any], *keys: str) -> float | None:
    for key in keys:
        parsed = _to_float(comp.get(key))
        if parsed is not None:
            return parsed
    return None


def _clear_comp_row(comps_ws: Worksheet, row: int) -> None:
    for col in ("B", "C", "D", "E", "G", "M", "O", "P", "Q", "R", "S"):
        _safe_set_or_clear(comps_ws, f"{col}{row}", None)


def _sanitize_growth_rate(value: float | None) -> float | None:
    if value is None:
        return None
    return max(-0.25, min(0.60, value))


def _normalized_ntm_metric(
    *,
    ltm_value: float | None,
    ntm_value: float | None,
    growth_rate: float | None = None,
    default_growth_rate: float = 0.03,
) -> float | None:
    # Prefer explicit NTM values when available.
    # If absent/invalid, build a forward proxy from free-source growth fields.
    if ntm_value is None or ntm_value <= 0:
        if ltm_value is None or ltm_value <= 0:
            return None
        growth = _sanitize_growth_rate(growth_rate)
        if growth is None:
            growth = _sanitize_growth_rate(default_growth_rate) or 0.03
        return ltm_value * (1.0 + growth)

    # Heuristic: payload may provide NTM in billions while LTM is in either
    # dollars or millions depending on upstream source.
    if ltm_value is not None and 0 < ntm_value < 1000:
        if ltm_value >= 1_000_000:
            # LTM likely in raw dollars; convert "billions" to dollars.
            return ntm_value * 1_000_000_000
        if ltm_value >= 1000:
            # LTM likely in millions; convert "billions" to millions.
            return ntm_value * 1000
    return ntm_value


def _map_comps(comps_ws: Worksheet, payload: dict[str, Any], divisor: float) -> None:
    comps = payload.get("comps", []) or []
    assumptions = payload.get("assumptions")
    assumptions = assumptions if isinstance(assumptions, dict) else {}
    default_growth_rate = _to_float(assumptions.get("revenueGrowth"))
    if default_growth_rate is None:
        default_growth_rate = _to_float(assumptions.get("revenueGrowthRate"))
    if default_growth_rate is None:
        default_growth_rate = 0.03

    for index in range(6):
        row = 8 + index
        comp = comps[index] if index < len(comps) and isinstance(comps[index], dict) else None
        if comp is None:
            _clear_comp_row(comps_ws, row)
            continue

        name = _normalize_comp_name(comp)
        price = _first_float(comp, "price", "sharePrice")
        shares = _normalize_comp_shares(_first_float(comp, "sharesOutstanding", "shares", "shares_outstanding"))
        beta = _first_float(comp, "beta")
        ev = _first_float(comp, "ev", "enterpriseValue", "enterprise_value")
        de_ratio = _first_float(comp, "debtToEquity", "deRatio", "de_ratio")
        ebitda = _first_float(comp, "ebitda", "ebitdaLtm", "ebitda_ltm")
        revenue = _first_float(comp, "revenue", "revenueLtm", "revenue_ltm")
        ebitda_ntm_raw = _first_float(comp, "ebitdaNtm", "ntmEbitda", "ebitda_ntm")
        revenue_ntm_raw = _first_float(comp, "revenueNtm", "ntmRevenue", "revenue_ntm")
        comp_growth = _first_float(comp, "revenueGrowth", "growth")
        ticker_raw = comp.get("ticker") or comp.get("symbol")
        ticker = ticker_raw.strip().upper() if isinstance(ticker_raw, str) and ticker_raw.strip() else None

        market_cap = _first_float(comp, "marketCap", "market_cap", "equityValue")
        total_debt = _first_float(comp, "totalDebt", "debt")
        cash = _first_float(comp, "cash")

        if ev is None and market_cap is not None:
            ev = market_cap + (total_debt or 0.0) - (cash or 0.0)
        if shares is None and market_cap is not None and price is not None and price > 0:
            shares = market_cap / price
        if ebitda is None:
            ev_ebitda = _first_float(comp, "evEbitda", "ev_ebitda")
            if ev is not None and ev_ebitda is not None and ev_ebitda > 0:
                ebitda = ev / ev_ebitda
        if revenue is None:
            ev_rev = _first_float(comp, "evRev", "evRevenue", "ev_rev", "ev_revenue")
            if ev is not None and ev_rev is not None and ev_rev > 0:
                revenue = ev / ev_rev
        if de_ratio is None and total_debt is not None and market_cap and market_cap > 0:
            de_ratio = total_debt / market_cap
        revenue_ntm = _normalized_ntm_metric(
            ltm_value=revenue,
            ntm_value=revenue_ntm_raw,
            growth_rate=comp_growth,
            default_growth_rate=default_growth_rate,
        )
        ebitda_ntm = _normalized_ntm_metric(
            ltm_value=ebitda,
            ntm_value=ebitda_ntm_raw,
            growth_rate=comp_growth,
            default_growth_rate=default_growth_rate,
        )

        _safe_set_or_clear(comps_ws, f"B{row}", name)
        _safe_set_or_clear(comps_ws, f"C{row}", price)
        _safe_set_or_clear(comps_ws, f"D{row}", shares)
        _safe_set_or_clear(comps_ws, f"E{row}", beta)
        _safe_set_or_clear(comps_ws, f"G{row}", _scale(ev, divisor) if ev is not None else None)
        _safe_set_or_clear(comps_ws, f"M{row}", de_ratio)
        _safe_set_or_clear(comps_ws, f"O{row}", ticker)
        _safe_set_or_clear(comps_ws, f"P{row}", _scale(ebitda, divisor) if ebitda is not None else None)
        _safe_set_or_clear(comps_ws, f"Q{row}", _scale(ebitda_ntm, divisor) if ebitda_ntm is not None else None)
        _safe_set_or_clear(comps_ws, f"R{row}", _scale(revenue, divisor) if revenue is not None else None)
        _safe_set_or_clear(comps_ws, f"S{row}", _scale(revenue_ntm, divisor) if revenue_ntm is not None else None)


def _build_timeline(payload: dict[str, Any]) -> tuple[list[int], set[int]]:
    historical_years = sorted({int(y) for y in (payload.get("historicals", {}).get("years", []) or []) if _to_float(y) is not None})
    forecast_years = sorted(
        {
            int(year)
            for year in [
                (_to_float(f.get("year")) if isinstance(f, dict) else None)
                for f in (payload.get("forecasts", []) or [])
            ]
            if year is not None
        }
    )

    combined = sorted({*historical_years, *forecast_years})
    if not combined:
        current_year = datetime.now().year
        combined = [current_year - 4 + i for i in range(10)]

    if len(combined) > 10:
        timeline = combined[-10:]
    else:
        timeline = combined[:]

    while len(timeline) < 10:
        timeline.append(timeline[-1] + 1)

    return timeline, set(historical_years)


def _map_year_headers(
    outputs: Worksheet,
    dcf_base: Worksheet,
    dcf_bull: Worksheet,
    dcf_bear: Worksheet,
    data_original: Worksheet,
    data_recalc: Worksheet,
    timeline_years: list[int],
    historical_years: set[int],
    payload: dict[str, Any],
) -> None:
    company = payload.get("company", {})
    as_of = _safe_date(company.get("asOfDate"))
    fallback_year = as_of.year if as_of is not None else datetime.now().year
    fiscal_end = _fiscal_year_end_date(company.get("fiscalYearEnd"), fallback_year) or date(fallback_year, 12, 31)
    fiscal_month = fiscal_end.month
    fiscal_day = fiscal_end.day

    # Force explicit FY labels so Actual/Forecast split is payload-driven.
    for idx, col in enumerate(DCF_TIMELINE_COLUMNS):
        year = timeline_years[idx]
        label = f"FY{year}{'A' if year in historical_years else 'E'}"
        _force_set(outputs, f"{col}6", label)
        _force_set(dcf_base, f"{col}18", label)
        _force_set(dcf_bull, f"{col}18", label)
        _force_set(dcf_bear, f"{col}18", label)
        _force_set(dcf_base, f"{col}63", label)
        _force_set(dcf_bull, f"{col}63", label)
        _force_set(dcf_bear, f"{col}63", label)
        _force_set(dcf_base, f"{col}72", label)
        _force_set(dcf_bull, f"{col}72", label)
        _force_set(dcf_bear, f"{col}72", label)

    assumptions = payload.get("assumptions", {})
    assumptions = assumptions if isinstance(assumptions, dict) else {}
    terminal_assumptions = assumptions.get("terminal")
    terminal_assumptions = terminal_assumptions if isinstance(terminal_assumptions, dict) else {}
    wacc_assumptions = assumptions.get("wacc")
    wacc_assumptions = wacc_assumptions if isinstance(wacc_assumptions, dict) else {}
    base_wacc = _sanitize_wacc_rate(assumptions.get("waccRate") or wacc_assumptions.get("waccRate"))
    _safe_set(outputs, "H28", _sanitize_terminal_growth_rate(terminal_assumptions.get("g"), reference_wacc=base_wacc))

    for idx, col in enumerate(TEN_YEAR_COLUMNS):
        year = timeline_years[idx]
        _safe_set(data_original, f"{col}3", "Actual" if year in historical_years else "Projections")
        _safe_set(data_original, f"{col}5", _safe_year_end_date(year, fiscal_month, fiscal_day))
        year = timeline_years[idx]
        _force_set(data_original, f"{col}6", f"FY{year}{'A' if year in historical_years else 'E'}")

    for idx, col in enumerate(RECALC_COLUMNS):
        year = timeline_years[idx]
        _safe_set(data_recalc, f"{col}3", "Actual" if year in historical_years else "Projections")
        _safe_set(data_recalc, f"{col}5", _safe_year_end_date(year, fiscal_month, fiscal_day))
        year = timeline_years[idx]
        _force_set(data_recalc, f"{col}6", f"FY{year}{'A' if year in historical_years else 'E'}")


def _map_data_sheets(
    data_original: Worksheet,
    data_recalc: Worksheet,
    payload: dict[str, Any],
    divisor: float,
    timeline_years: list[int],
) -> None:
    revenue_series = _metric_series(payload, timeline_years, "revenue")
    cost_of_revenue_series = _metric_series(payload, timeline_years, "cost_of_revenue", revenue_series)
    sales_commission_series = _metric_series(payload, timeline_years, "sales_commission", revenue_series)
    purchases_series, sales_commission_series = _split_cost_of_revenue_components(
        cost_of_revenue_series,
        sales_commission_series,
    )
    opex_components = _opex_component_series(
        data_original,
        payload,
        timeline_years,
        revenue_series,
        purchases_series,
        sales_commission_series,
    )

    for idx, column in enumerate(TEN_YEAR_COLUMNS):
        _force_set(data_original, f"{column}12", _scale(revenue_series[idx], divisor))
        _force_set(data_original, f"{column}16", _scale(purchases_series[idx], divisor))
        _force_set(data_original, f"{column}17", _scale(sales_commission_series[idx], divisor))
        _force_set(data_original, f"{column}24", _scale(opex_components["rnd"][idx], divisor))
        _force_set(data_original, f"{column}25", _scale(opex_components["sga"][idx], divisor))
        _force_set(data_original, f"{column}26", _scale(opex_components["da"][idx], divisor))
        _force_set(data_original, f"{column}27", _scale(opex_components["other"][idx], divisor))

    for idx, column in enumerate(RECALC_COLUMNS):
        _force_set(data_recalc, f"{column}12", _scale(revenue_series[idx], divisor))
        _force_set(data_recalc, f"{column}16", _scale(purchases_series[idx], divisor))
        _force_set(data_recalc, f"{column}17", _scale(sales_commission_series[idx], divisor))
        _force_set(data_recalc, f"{column}24", _scale(opex_components["rnd"][idx], divisor))
        _force_set(data_recalc, f"{column}25", _scale(opex_components["sga"][idx], divisor))
        _force_set(data_recalc, f"{column}26", _scale(opex_components["da"][idx], divisor))
        _force_set(data_recalc, f"{column}27", _scale(opex_components["other"][idx], divisor))

    # Remove legacy M&A assumption block from both data tabs.
    for sheet in (data_original, data_recalc):
        for row in range(34, 42):
            for col in ("B", "C"):
                _safe_set_or_clear(sheet, f"{col}{row}", None)
    _safe_set_or_clear(data_recalc, "R30", None)


def _historical_value_for_year(
    payload: dict[str, Any],
    year: int,
    *,
    statement: str,
    keys: list[str],
) -> float | None:
    historicals = payload.get("historicals", {})
    if not isinstance(historicals, dict):
        return None
    section = historicals.get(statement)
    if not isinstance(section, dict):
        return None
    values = _series(section, keys)
    index = _historical_index_by_year(payload).get(year)
    if index is None or index >= len(values):
        return None
    value = _to_float(values[index])
    return abs(value) if value is not None else None


def _add_prior_actual_year_display_column(
    dcf_base: Worksheet,
    dcf_bull: Worksheet,
    dcf_bear: Worksheet,
    payload: dict[str, Any],
    timeline_years: list[int],
    historical_years: set[int],
    divisor: float,
) -> None:
    if not timeline_years:
        return
    prior_candidates = [year for year in historical_years if year < timeline_years[0]]
    if not prior_candidates:
        return

    prior_year = max(prior_candidates)
    prior_label = f"FY{prior_year}A"

    revenue = _historical_value_for_year(payload, prior_year, statement="income", keys=["Total Revenue", "Revenue", "Sales"])
    cost_of_revenue = _historical_value_for_year(
        payload,
        prior_year,
        statement="income",
        keys=["Cost of Revenue", "COGS", "Cost Of Revenue", "Cost of Sales", "Purchases"],
    )
    rnd = _historical_value_for_year(payload, prior_year, statement="income", keys=["Research & Development", "R&D", "Research and Development"])
    sga = _historical_value_for_year(
        payload,
        prior_year,
        statement="income",
        keys=["SG&A", "SGA", "General and Administrative", "GeneralAndAdministrative", "G&A", "GA"],
    )
    da = _historical_value_for_year(
        payload,
        prior_year,
        statement="income",
        keys=["D&A (included in Operating)", "D&A", "DA", "Depreciation & Amortization", "Depreciation"],
    )
    if da is None:
        da = _historical_value_for_year(payload, prior_year, statement="cashflow", keys=["Depreciation"])
    other_opex = _historical_value_for_year(payload, prior_year, statement="income", keys=["Other Operating Expenses", "Other"])
    ebit = _historical_value_for_year(payload, prior_year, statement="income", keys=["Operating Income (EBIT)", "EBIT", "Operating Income"])
    capex = _historical_value_for_year(payload, prior_year, statement="cashflow", keys=["Capex", "Capital Expenditures", "Capital Expenditure"])

    for sheet in (dcf_base, dcf_bull, dcf_bear):
        # Carry timeline header style to the added prior-year display column.
        for row in (18, 63, 72):
            sheet[f"G{row}"]._style = copy(sheet[f"H{row}"]._style)
            _force_set(sheet, f"G{row}", prior_label)

        # Apply consistent number/border styles from first timeline column.
        for row in (20, 24, 27, 30, 32, 33, 36, 39, 42, 45, 48, 49, 51, 52, 54, 55, 57, 60, 65, 66, 67, 68, 69, 74, 75, 76, 77, 78, 79):
            sheet[f"G{row}"]._style = copy(sheet[f"H{row}"]._style)

        if revenue is not None:
            _force_set(sheet, "G20", _scale(revenue, divisor))
            _force_set(sheet, "H21", "=IFERROR(H20/G20-1,0)")
        _force_set(sheet, "G21", "-")
        if cost_of_revenue is not None:
            _force_set(sheet, "G24", _scale(cost_of_revenue, divisor))
        _force_set(sheet, "G25", "=IFERROR(G24/G20,0)")
        _force_set(sheet, "G28", "=IFERROR(G27/G20,0)")
        _force_set(sheet, "G30", "=G24+G27")
        _force_set(sheet, "G32", "=G20-G30")
        _force_set(sheet, "G33", "=IFERROR(G32/G20,0)")
        if rnd is not None:
            _force_set(sheet, "G36", _scale(rnd, divisor))
        _force_set(sheet, "G37", "=IFERROR(G36/G20,0)")
        if sga is not None:
            _force_set(sheet, "G39", _scale(sga, divisor))
        _force_set(sheet, "G40", "=IFERROR(G39/G20,0)")
        if da is not None:
            _force_set(sheet, "G42", _scale(da, divisor))
        _force_set(sheet, "G43", "=IFERROR(G42/G20,0)")
        if other_opex is not None:
            _force_set(sheet, "G45", _scale(other_opex, divisor))
        _force_set(sheet, "G46", "=IFERROR(G45/G20,0)")
        _force_set(sheet, "G48", "=G36+G39+G42+G45")
        _force_set(sheet, "G49", "=IFERROR(G48/G20,0)")
        if ebit is not None:
            _force_set(sheet, "G51", _scale(ebit, divisor))
        else:
            _force_set(sheet, "G51", "=G32-G48")
        _force_set(sheet, "G52", "=IFERROR(G51/G20,0)")
        _force_set(sheet, "G54", "=G51+G68")
        _force_set(sheet, "G55", "=IFERROR(G54/G20,0)")
        _force_set(sheet, "G57", "=-G51*$F$11")
        _force_set(sheet, "G58", "=$F$11")
        _force_set(sheet, "G60", "=G51+G57")
        _force_set(sheet, "G61", "=IFERROR(G60/G20,0)")

        if capex is not None:
            _force_set(sheet, "G65", -(_scale(capex, divisor) or 0.0))
        else:
            _force_set(sheet, "G65", "=H65")
        _force_set(sheet, "G66", "=IFERROR(-G65/G20,0)")
        _safe_set_or_clear(sheet, "G67", None)
        if da is not None:
            _force_set(sheet, "G68", _scale(da, divisor))
        else:
            _force_set(sheet, "G68", "=G69*G20")
        _force_set(sheet, "G69", "=$F$13")
        _force_set(sheet, "G70", "=IFERROR(-G68/G65,0)")

        _force_set(sheet, "G74", "=G60")
        _force_set(sheet, "G75", "=G68")
        _force_set(sheet, "G76", "=G65")
        _force_set(sheet, "G77", "=0")
        _force_set(sheet, "G78", "=SUM(G74:G77)")
        _force_set(sheet, "G79", "=IFERROR(G78/G20,0)")

def _historical_index_by_year(payload: dict[str, Any]) -> dict[int, int]:
    years = payload.get("historicals", {}).get("years", []) or []
    return {int(year): idx for idx, year in enumerate(years) if _to_float(year) is not None}


def _forecast_by_year(payload: dict[str, Any]) -> dict[int, dict[str, Any]]:
    mapping: dict[int, dict[str, Any]] = {}
    for forecast in payload.get("forecasts", []) or []:
        if not isinstance(forecast, dict):
            continue
        year = _to_float(forecast.get("year"))
        if year is None:
            continue
        mapping[int(year)] = forecast
    return mapping


def _metric_series(payload: dict[str, Any], timeline: list[int], metric: str, revenue_series: list[float] | None = None) -> list[float]:
    historicals = payload.get("historicals", {})
    income = historicals.get("income", {}) if isinstance(historicals, dict) else {}

    mapping = {
        "revenue": ["Total Revenue", "Revenue", "Sales"],
        "cost_of_revenue": ["Cost of Revenue", "COGS", "Cost Of Revenue", "Cost of Sales", "Purchases"],
        "sales_commission": ["Sales Commission", "SalesCommission"],
    }
    source = _series(income, mapping[metric])

    idx_by_year = _historical_index_by_year(payload)
    forecast_by_year = _forecast_by_year(payload)

    series: list[float] = []
    fallback_ratio = 0.0
    if metric == "cost_of_revenue":
        fallback_ratio = _last_known_ratio(source, revenue_series or [], default_ratio=0.60)
    elif metric == "sales_commission":
        fallback_ratio = _last_known_ratio(source, revenue_series or [], default_ratio=0.0)

    for idx, year in enumerate(timeline):
        forecast = forecast_by_year.get(year)
        value = None
        if forecast:
            if metric == "revenue":
                value = _to_float(forecast.get("revenue"))
            elif metric == "cost_of_revenue":
                value = _first_float(
                    forecast,
                    "costOfRevenue",
                    "cogs",
                    "cost_of_revenue",
                    "costOfSales",
                    "cost_of_sales",
                )
                if (value is None or value <= 0) and revenue_series is not None and idx < len(revenue_series):
                    rev = revenue_series[idx]
                    if rev > 0:
                        value = rev * fallback_ratio
            else:
                value = _first_float(
                    forecast,
                    "salesCommission",
                    "sales_commission",
                    "marketingExpense",
                    "marketing",
                )
                rev = revenue_series[idx] if revenue_series is not None and idx < len(revenue_series) else _to_float(forecast.get("revenue"))
                if value is not None and value > 0:
                    pass
                elif rev is not None and rev > 0:
                    value = rev * fallback_ratio

        hist_idx = idx_by_year.get(year)
        is_projection = hist_idx is None
        if hist_idx is not None and hist_idx < len(source):
            hist_value = abs(source[hist_idx])
            if hist_value > 0:
                series.append(hist_value)
                continue

            if metric in {"cost_of_revenue", "sales_commission"} and is_projection:
                rev = revenue_series[idx] if revenue_series is not None and idx < len(revenue_series) else 0.0
                if rev > 0:
                    series.append(rev * fallback_ratio)
                    continue

        if value is not None:
            value_abs = abs(value)
            if metric in {"cost_of_revenue", "sales_commission"} and is_projection and value_abs <= 0:
                rev = revenue_series[idx] if revenue_series is not None and idx < len(revenue_series) else 0.0
                if rev > 0:
                    value_abs = rev * fallback_ratio
            series.append(value_abs)
            continue

        previous = series[-1] if series else 0.0
        series.append(previous)

    return series


def _capex_series(payload: dict[str, Any], timeline: list[int]) -> list[float]:
    historicals = payload.get("historicals", {})
    cashflow = historicals.get("cashflow", {}) if isinstance(historicals, dict) else {}
    source = _series(cashflow, ["Capex", "Capital Expenditures", "Capital Expenditure"])

    assumptions = payload.get("assumptions", {})
    default_capex = None
    if isinstance(assumptions, dict):
        capex_abs = assumptions.get("capexAbsolute")
        if isinstance(capex_abs, list) and capex_abs:
            default_capex = _to_float(capex_abs[0])

    forecast_by_year = _forecast_by_year(payload)
    first_forecast = payload.get("forecasts", [None])[0]
    if default_capex is None and isinstance(first_forecast, dict):
        default_capex = _to_float(first_forecast.get("capex"))
    default_capex = abs(default_capex) if default_capex is not None else 0.0

    idx_by_year = _historical_index_by_year(payload)
    series: list[float] = []
    for year in timeline:
        hist_idx = idx_by_year.get(year)
        value = None
        if hist_idx is not None and hist_idx < len(source):
            historical_capex = abs(source[hist_idx])
            if historical_capex > 0:
                value = historical_capex

        if value is None:
            forecast = forecast_by_year.get(year)
            if isinstance(forecast, dict):
                forecast_capex = _to_float(forecast.get("capex"))
                if forecast_capex is not None:
                    value = abs(forecast_capex)

        if value is None:
            value = series[-1] if series else default_capex

        series.append(value)

    return series

def _split_cost_of_revenue_components(
    cost_of_revenue_series: list[float],
    sales_commission_series: list[float],
) -> tuple[list[float], list[float]]:
    purchases_series: list[float] = []
    normalized_sales_commission: list[float] = []
    total_periods = max(len(cost_of_revenue_series), len(sales_commission_series))
    for idx in range(total_periods):
        cogs = abs(cost_of_revenue_series[idx]) if idx < len(cost_of_revenue_series) else 0.0
        sales_commission = abs(sales_commission_series[idx]) if idx < len(sales_commission_series) else 0.0
        sales_commission = min(sales_commission, cogs)
        purchases_series.append(max(0.0, cogs - sales_commission))
        normalized_sales_commission.append(sales_commission)
    return purchases_series, normalized_sales_commission


def _last_known_ratio(numerator_series: list[float], denominator_series: list[float], *, default_ratio: float) -> float:
    size = min(len(numerator_series), len(denominator_series))
    for idx in range(size - 1, -1, -1):
        denominator = denominator_series[idx]
        numerator = numerator_series[idx]
        if denominator and numerator:
            return max(0.0, min(1.0, numerator / denominator))
    return max(0.0, min(1.0, default_ratio))


def _opex_component_series(
    template_sheet: Worksheet,
    payload: dict[str, Any],
    timeline: list[int],
    revenue_series: list[float],
    purchases_series: list[float],
    sales_commission_series: list[float],
) -> dict[str, list[float]]:
    historicals = payload.get("historicals", {})
    income = historicals.get("income", {}) if isinstance(historicals, dict) else {}
    cashflow = historicals.get("cashflow", {}) if isinstance(historicals, dict) else {}

    # Prefer website line items; keep legacy keys as fallback.
    rnd_series_hist = _series(income, ["Research & Development", "R&D", "Research and Development"])
    sga_series_hist = _series(
        income,
        ["SG&A", "SGA", "General and Administrative", "GeneralAndAdministrative", "G&A", "GA"],
    )
    da_series_hist = _series(
        income,
        ["D&A (included in Operating)", "D&A", "DA", "Depreciation & Amortization", "Depreciation"],
    )
    if not da_series_hist:
        da_series_hist = _series(cashflow, ["Depreciation"])
    other_series_hist = _series(income, ["Other Operating Expenses", "Other"])
    operating_exp_hist = _series(income, ["Operating Expenses", "Total Operating Expenses", "OperatingExpense", "OPEX"])
    ebit_hist = _series(income, ["Operating Income (EBIT)", "EBIT", "Operating Income"])
    gross_profit_hist = _series(income, ["Gross Profit", "GrossProfit"])
    revenue_hist = _series(income, ["Total Revenue", "Revenue", "Sales"])

    idx_by_year = _historical_index_by_year(payload)
    forecast_by_year = _forecast_by_year(payload)

    mix = _last_known_opex_mix(
        rnd_series_hist,
        sga_series_hist,
        da_series_hist,
        other_series_hist,
    ) or _template_opex_mix(template_sheet)

    out = {"rnd": [], "sga": [], "da": [], "other": []}
    for idx, year in enumerate(timeline):
        hist_idx = idx_by_year.get(year)
        forecast = forecast_by_year.get(year)
        revenue = _to_float(forecast.get("revenue")) if forecast else None
        ebit = _to_float(forecast.get("ebit")) if forecast else None
        gross_profit = _to_float(forecast.get("grossProfit")) if forecast else None

        if revenue is None and hist_idx is not None and hist_idx < len(revenue_hist):
            revenue = revenue_hist[hist_idx]
        if revenue is None and idx < len(revenue_series):
            revenue = revenue_series[idx]
        if ebit is None and hist_idx is not None and hist_idx < len(ebit_hist):
            ebit = ebit_hist[hist_idx]
        if gross_profit is None and hist_idx is not None and hist_idx < len(gross_profit_hist):
            gross_profit = gross_profit_hist[hist_idx]

        operating_total = None
        if hist_idx is not None and hist_idx < len(operating_exp_hist):
            operating_total = abs(operating_exp_hist[hist_idx])
        if operating_total is None and gross_profit is not None and ebit is not None:
            operating_total = max(0.0, gross_profit - ebit)
        if operating_total is None and revenue is not None and ebit is not None:
            operating_total = max(0.0, revenue - purchases_series[idx] - sales_commission_series[idx] - ebit)
        if operating_total is None:
            operating_total = (
                (out["rnd"][-1] if out["rnd"] else 0.0)
                + (out["sga"][-1] if out["sga"] else 0.0)
                + (out["da"][-1] if out["da"] else 0.0)
                + (out["other"][-1] if out["other"] else 0.0)
            )

        rnd = (
            abs(rnd_series_hist[hist_idx])
            if hist_idx is not None and hist_idx < len(rnd_series_hist) and abs(rnd_series_hist[hist_idx]) > 0
            else abs(_to_float(forecast.get("rdExpense")) or 0.0) if forecast else 0.0
        )
        sga = (
            abs(sga_series_hist[hist_idx])
            if hist_idx is not None and hist_idx < len(sga_series_hist) and abs(sga_series_hist[hist_idx]) > 0
            else abs(_to_float(forecast.get("sgaExpense")) or 0.0) if forecast else 0.0
        )
        da = (
            abs(da_series_hist[hist_idx])
            if hist_idx is not None and hist_idx < len(da_series_hist) and abs(da_series_hist[hist_idx]) > 0
            else abs(_to_float(forecast.get("depreciation")) or 0.0) if forecast else 0.0
        )
        other = abs(other_series_hist[hist_idx]) if hist_idx is not None and hist_idx < len(other_series_hist) else 0.0

        provided = {"rnd": rnd, "sga": sga, "da": da, "other": other}
        missing = [k for k, v in provided.items() if v <= 0]
        known_sum = sum(v for v in provided.values() if v > 0)
        remaining = max(0.0, operating_total - known_sum)

        if missing:
            missing_mix_total = sum(mix[key] for key in missing)
            for key in missing:
                weight = (mix[key] / missing_mix_total) if missing_mix_total > 0 else (1.0 / len(missing))
                provided[key] = remaining * weight
        elif operating_total > known_sum and known_sum > 0:
            # Keep historical line items and push reconciliation into "other".
            provided["other"] = max(0.0, provided["other"] + (operating_total - known_sum))

        out["rnd"].append(provided["rnd"])
        out["sga"].append(provided["sga"])
        out["da"].append(provided["da"])
        out["other"].append(provided["other"])

    return out


def _last_known_opex_mix(
    rnd_series: list[float],
    sga_series: list[float],
    da_series: list[float],
    other_series: list[float],
) -> dict[str, float] | None:
    size = min(len(rnd_series), len(sga_series), len(da_series), len(other_series))
    for idx in range(size - 1, -1, -1):
        rnd = abs(rnd_series[idx])
        sga = abs(sga_series[idx])
        da = abs(da_series[idx])
        other = abs(other_series[idx])
        total = rnd + sga + da + other
        if total > 0:
            return {
                "rnd": rnd / total,
                "sga": sga / total,
                "da": da / total,
                "other": other / total,
            }
    return None


def _template_opex_mix(template_sheet: Worksheet) -> dict[str, float]:
    totals = {
        "rnd": abs(_to_float(template_sheet["X24"].value) or 0.0),
        "sga": abs(_to_float(template_sheet["X25"].value) or 0.0),
        "da": abs(_to_float(template_sheet["X26"].value) or 0.0),
        "other": abs(_to_float(template_sheet["X27"].value) or 0.0),
    }
    total = sum(totals.values())
    if total <= 0:
        return {"rnd": 0.25, "sga": 0.50, "da": 0.25, "other": 0.0}
    return {
        "rnd": totals["rnd"] / total,
        "sga": totals["sga"] / total,
        "da": totals["da"] / total,
        "other": totals["other"] / total,
    }


def _remove_assumption_breakdown(workbook: Workbook, cover: Worksheet) -> None:
    if SHEET_ASSUMPTION_BREAKDOWN in workbook.sheetnames:
        workbook.remove(workbook[SHEET_ASSUMPTION_BREAKDOWN])
    if SHEET_DATA_ORIGINAL in workbook.sheetnames:
        workbook.remove(workbook[SHEET_DATA_ORIGINAL])
    if "Data ->" in workbook.sheetnames:
        workbook.remove(workbook["Data ->"])

    # Keep cover TOC coherent after removing helper tabs.
    _safe_set(cover, "F15", "Data Given (Recalculated)")
    _safe_set_or_clear(cover, "F16", None)
    _safe_set_or_clear(cover, "F17", None)


def _replace_template_placeholders(
    *,
    company_name: str | None,
    ticker: str,
    sheets: tuple[Worksheet, ...],
) -> None:
    long_name = company_name or ticker
    replacements = {
        "ABC/SNS Health": long_name,
        "ABC Corp.": long_name,
        "ABC Corp": long_name,
        "ABC": ticker,
    }

    for sheet in sheets:
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                if cell.data_type == "f":
                    continue
                if not isinstance(cell.value, str):
                    continue

                updated = cell.value
                for old, new in replacements.items():
                    updated = updated.replace(old, new)
                if updated != cell.value:
                    cell.value = updated
