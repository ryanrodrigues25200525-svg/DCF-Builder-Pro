from __future__ import annotations

import re
from copy import copy, deepcopy
from functools import lru_cache
from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile

from defusedxml import ElementTree as ET
from lxml import etree as LET
from openpyxl import load_workbook

from app.services.excel_export.mappers import (
    WACC_LOOP_MODE_ITERATIVE,
    apply_payload_to_workbook,
    resolve_wacc_loop_mode,
)
from app.services.excel_export.template import load_template_artifact

_TEMPLATE_SHEET_NAME_ALIASES = {
    "Outputs - Base": "Ouputs - Base",
}

_DCF_SCENARIO_SHEET_NAMES = (
    "DCF Model - Base (1)",
    "DCF Model - Bull (2)",
    "DCF Model - Bear (3)",
)

_NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

_STYLE_SOURCE_OVERRIDES: dict[str, dict[str, str]] = {
    "Outputs - Base": {
        "I6": "H6",
        "J6": "H6",
        "K6": "H6",
        "L6": "H6",
        "M6": "H6",
        "N6": "H6",
        "O6": "H6",
        "P6": "H6",
        "Q6": "H6",
    },
    "WACC": {
        "D35": "D34",
        "D36": "D34",
        "B37": "B23",
        "B38": "B23",
        "B39": "B23",
        "B40": "B34",
        "D37": "D11",
        "D38": "D11",
        "D39": "D16",
        "D40": "D34",
        "C40": "C34",
    },
    "DCF Model - Base (1)": {
        "C13": "F11",
        "C14": "F11",
        "C17": "C11",
        "D119": "C120",
        "E119": "C120",
        "F119": "C120",
        "G119": "C120",
        "H119": "C120",
        "J119": "C120",
        "K119": "C120",
        "L119": "C120",
        "M119": "C120",
        "N119": "C120",
        "I120": "C120",
        "I121": "C120",
        "I122": "C120",
        "I123": "C120",
        "I124": "C120",
        "E15": "D15",
        "F16": "D16",
        "F14": "F11",
        "G18": "H18",
        "G20": "H20",
        "G21": "H21",
        "G24": "H24",
        "G27": "H27",
        "G30": "H30",
        "G32": "H32",
        "G33": "H33",
        "G36": "H36",
        "G39": "H39",
        "G42": "H42",
        "G45": "H45",
        "G48": "H48",
        "G49": "H49",
        "G51": "H51",
        "G52": "H52",
        "G54": "H54",
        "G55": "H55",
        "G57": "H57",
        "G60": "H60",
        "G63": "H63",
        "G65": "H65",
        "G66": "H66",
        "G67": "H67",
        "G68": "H68",
        "G69": "H69",
        "G72": "H72",
        "G74": "H74",
        "G75": "H75",
        "G76": "H76",
        "G77": "H77",
        "G78": "H78",
        "G79": "H79",
        "H18": "I18",
        "H79": "J79",
        "I79": "J79",
        "H63": "H18",
        "I63": "I18",
        "J63": "J18",
        "K63": "K18",
        "L63": "L18",
        "M63": "M18",
        "N63": "N18",
        "O63": "O18",
        "P63": "P18",
        "Q63": "Q18",
        "H72": "H18",
        "I72": "I18",
        "J72": "J18",
        "K72": "K18",
        "L72": "L18",
        "M72": "M18",
        "N72": "N18",
        "O72": "O18",
        "P72": "P18",
        "Q72": "Q18",
        "J126": "C126",
        "J127": "C127",
        "J128": "C128",
        "J129": "C129",
        "J130": "C130",
        "I117": "C117",
        "J118": "D118",
        "I121": "B121",
        "I122": "B122",
    },
    "DCF Model - Bull (2)": {
        "C13": "F11",
        "C14": "F11",
        "C17": "C11",
        "E15": "D15",
        "F16": "D16",
        "F14": "F11",
        "G18": "H18",
        "G20": "H20",
        "G21": "H21",
        "G24": "H24",
        "G27": "H27",
        "G30": "H30",
        "G32": "H32",
        "G33": "H33",
        "G36": "H36",
        "G39": "H39",
        "G42": "H42",
        "G45": "H45",
        "G48": "H48",
        "G49": "H49",
        "G51": "H51",
        "G52": "H52",
        "G54": "H54",
        "G55": "H55",
        "G57": "H57",
        "G60": "H60",
        "G63": "H63",
        "G65": "H65",
        "G66": "H66",
        "G67": "H67",
        "G68": "H68",
        "G69": "H69",
        "G72": "H72",
        "G74": "H74",
        "G75": "H75",
        "G76": "H76",
        "G77": "H77",
        "G78": "H78",
        "G79": "H79",
        "H18": "I18",
        "H79": "J79",
        "I79": "J79",
        "H63": "H18",
        "I63": "I18",
        "J63": "J18",
        "K63": "K18",
        "L63": "L18",
        "M63": "M18",
        "N63": "N18",
        "O63": "O18",
        "P63": "P18",
        "Q63": "Q18",
        "H72": "H18",
        "I72": "I18",
        "J72": "J18",
        "K72": "K18",
        "L72": "L18",
        "M72": "M18",
        "N72": "N18",
        "O72": "O18",
        "P72": "P18",
        "Q72": "Q18",
        "J126": "C126",
        "J127": "C127",
        "J128": "C128",
        "J129": "C129",
        "J130": "C130",
        "I117": "C117",
        "J118": "D118",
        "I121": "B121",
        "I122": "B122",
    },
    "DCF Model - Bear (3)": {
        "C13": "F11",
        "C14": "F11",
        "C17": "C11",
        "E15": "D15",
        "F16": "D16",
        "F14": "F11",
        "G18": "H18",
        "G20": "H20",
        "G21": "H21",
        "G24": "H24",
        "G27": "H27",
        "G30": "H30",
        "G32": "H32",
        "G33": "H33",
        "G36": "H36",
        "G39": "H39",
        "G42": "H42",
        "G45": "H45",
        "G48": "H48",
        "G49": "H49",
        "G51": "H51",
        "G52": "H52",
        "G54": "H54",
        "G55": "H55",
        "G57": "H57",
        "G60": "H60",
        "G63": "H63",
        "G65": "H65",
        "G66": "H66",
        "G67": "H67",
        "G68": "H68",
        "G69": "H69",
        "G72": "H72",
        "G74": "H74",
        "G75": "H75",
        "G76": "H76",
        "G77": "H77",
        "G78": "H78",
        "G79": "H79",
        "H18": "I18",
        "H79": "J79",
        "I79": "J79",
        "H63": "H18",
        "I63": "I18",
        "J63": "J18",
        "K63": "K18",
        "L63": "L18",
        "M63": "M18",
        "N63": "N18",
        "O63": "O18",
        "P63": "P18",
        "Q63": "Q18",
        "H72": "H18",
        "I72": "I18",
        "J72": "J18",
        "K72": "K18",
        "L72": "L18",
        "M72": "M18",
        "N72": "N18",
        "O72": "O18",
        "P72": "P18",
        "Q72": "Q18",
        "J126": "C126",
        "J127": "C127",
        "J128": "C128",
        "J129": "C129",
        "J130": "C130",
        "I117": "C117",
        "J118": "D118",
        "I121": "B121",
        "I122": "B122",
    },
}


def export_dcf_excel(payload: dict) -> bytes:
    template = load_template_artifact()

    workbook = load_workbook(filename=BytesIO(template.workbook_bytes), data_only=False)
    apply_payload_to_workbook(workbook, payload)

    _apply_calculation_properties(workbook, payload)

    output_buffer = BytesIO()
    workbook.save(output_buffer)

    patched_output = _restore_template_styles(
        workbook_bytes=output_buffer.getvalue(),
        template_workbook_bytes=template.workbook_bytes,
        template_styles_xml=template.styles_xml,
    )
    return _finalize_output_year_labels(patched_output)


def _apply_calculation_properties(workbook, payload: dict) -> None:
    calc = workbook.calculation
    calc.fullCalcOnLoad = True
    calc.forceFullCalc = True

    loop_mode = resolve_wacc_loop_mode(payload)
    if loop_mode == WACC_LOOP_MODE_ITERATIVE:
        calc.iterate = True
        calc.iterateCount = 100
        calc.iterateDelta = 0.001
        return

    calc.iterate = False
    calc.iterateCount = None
    calc.iterateDelta = None


def _restore_template_styles(
    workbook_bytes: bytes,
    template_workbook_bytes: bytes,
    template_styles_xml: bytes,
) -> bytes:
    source = BytesIO(workbook_bytes)
    target = BytesIO()
    (
        patched_styles_xml,
        style_id_overrides_by_sheet,
        template_sheet_xml_by_name,
    ) = _cached_template_style_patch(template_workbook_bytes, template_styles_xml)

    with (
        ZipFile(source, "r") as in_zip,
        ZipFile(target, "w", compression=ZIP_DEFLATED) as out_zip,
    ):
        source_sheet_name_by_path = _sheet_name_by_path(in_zip)

        for item in in_zip.infolist():
            if item.filename == "xl/styles.xml":
                out_zip.writestr(item, patched_styles_xml)
                continue

            if item.filename.startswith("xl/worksheets/sheet"):
                source_name = source_sheet_name_by_path.get(item.filename)
                template_sheet_xml = None
                if source_name:
                    template_sheet_xml = template_sheet_xml_by_name.get(source_name)
                    if template_sheet_xml is None:
                        legacy_name = _TEMPLATE_SHEET_NAME_ALIASES.get(source_name)
                        if legacy_name:
                            template_sheet_xml = template_sheet_xml_by_name.get(legacy_name)
                if template_sheet_xml is not None:
                    patched_sheet = _patch_sheet_style_ids(
                        sheet_xml=in_zip.read(item.filename),
                        template_sheet_xml=template_sheet_xml,
                        sheet_name=source_name,
                        style_id_overrides=style_id_overrides_by_sheet.get(source_name),
                    )
                    out_zip.writestr(item, patched_sheet)
                    continue

            out_zip.writestr(item, in_zip.read(item.filename))

    return target.getvalue()


@lru_cache(maxsize=2)
def _cached_template_style_patch(
    template_workbook_bytes: bytes,
    template_styles_xml: bytes,
) -> tuple[bytes, dict[str, dict[str, str]], dict[str, bytes]]:
    with ZipFile(BytesIO(template_workbook_bytes), "r") as template_zip:
        template_sheet_name_by_path = _sheet_name_by_path(template_zip)
        template_sheet_path_by_name = {name: path for path, name in template_sheet_name_by_path.items()}
        template_cell_styles_by_sheet = _template_cell_styles_by_sheet(template_zip, template_sheet_path_by_name)
        patched_styles_xml, style_id_overrides_by_sheet = _patch_styles_xml(
            template_styles_xml,
            template_cell_styles_by_sheet,
        )
        template_sheet_xml_by_name: dict[str, bytes] = {}
        for sheet_name, sheet_path in template_sheet_path_by_name.items():
            if sheet_path in template_zip.namelist():
                template_sheet_xml_by_name[sheet_name] = template_zip.read(sheet_path)
        return patched_styles_xml, style_id_overrides_by_sheet, template_sheet_xml_by_name


def _sheet_name_by_path(archive: ZipFile) -> dict[str, str]:
    workbook_xml = ET.fromstring(archive.read("xl/workbook.xml"))
    rels_xml = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))

    ns_workbook = {
        "x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }
    ns_rels = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}

    rel_target_by_id: dict[str, str] = {}
    for rel in rels_xml.findall("r:Relationship", ns_rels):
        rel_id = rel.attrib.get("Id")
        target = rel.attrib.get("Target")
        if not rel_id or not target:
            continue
        rel_target_by_id[rel_id] = _normalize_rel_target(target)

    sheet_name_by_path: dict[str, str] = {}
    for sheet in workbook_xml.findall("x:sheets/x:sheet", ns_workbook):
        name = sheet.attrib.get("name")
        rel_id = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
        if not name or not rel_id:
            continue
        target = rel_target_by_id.get(rel_id)
        if target:
            sheet_name_by_path[target] = name

    return sheet_name_by_path


def _normalize_rel_target(target: str) -> str:
    normalized = target.strip().replace("\\", "/")
    if normalized.startswith("/"):
        normalized = normalized[1:]
    while normalized.startswith("./"):
        normalized = normalized[2:]
    if not normalized.startswith("xl/"):
        normalized = f"xl/{normalized}"
    return normalized


def _finalize_output_year_labels(workbook_bytes: bytes) -> bytes:
    # Final hard-guard to ensure output headers stay as explicit FY text labels.
    wb = load_workbook(filename=BytesIO(workbook_bytes), data_only=False)
    if "Outputs - Base" not in wb.sheetnames or "DCF Model - Base (1)" not in wb.sheetnames:
        out = BytesIO()
        wb.save(out)
        return out.getvalue()

    outputs = wb["Outputs - Base"]
    base = wb["DCF Model - Base (1)"]
    dcf_sheets = [wb[name] for name in ("DCF Model - Base (1)", "DCF Model - Bull (2)", "DCF Model - Bear (3)") if name in wb.sheetnames]
    for col in "HIJKLMNOPQ":
        label = base[f"{col}18"].value
        if not (isinstance(label, str) and label.startswith("FY")):
            continue
        outputs[f"{col}6"].value = label
        outputs[f"{col}6"].number_format = "@"
        for sheet in dcf_sheets:
            sheet[f"{col}89"].value = label
            sheet[f"{col}89"].number_format = "@"

    # Prior-year display column (G) must visually match first timeline column (H)
    # for heading, number format, font size, and borders.
    g_mirror_rows = (
        18, 20, 21, 24, 25, 27, 28, 30, 32, 33, 36, 37, 39, 40, 42, 43, 45, 46,
        48, 49, 51, 52, 54, 55, 57, 58, 60, 61, 63, 65, 66, 67, 68, 69, 70, 72,
        74, 75, 76, 77, 78, 79,
    )
    for sheet in dcf_sheets:
        if sheet["G18"].value in (None, ""):
            continue
        for row in g_mirror_rows:
            sheet[f"G{row}"]._style = copy(sheet[f"H{row}"]._style)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _template_cell_styles_by_sheet(
    template_zip: ZipFile,
    template_sheet_path_by_name: dict[str, str],
) -> dict[str, dict[str, str]]:
    ns = {"x": _NS_MAIN}
    out: dict[str, dict[str, str]] = {}
    for sheet_name, sheet_path in template_sheet_path_by_name.items():
        if sheet_path not in template_zip.namelist():
            continue
        root = ET.fromstring(template_zip.read(sheet_path))
        styles: dict[str, str] = {}
        for cell in root.findall(".//x:c", ns):
            address = cell.attrib.get("r")
            style_id = cell.attrib.get("s")
            if not address or style_id is None:
                continue
            styles[address] = style_id
        out[sheet_name] = styles
    return out


def _clone_cell_xf_with_font(
    cell_xfs,
    xfs,
    *,
    source_style_id: int,
    font_id: int,
) -> int:
    source_xf = xfs[source_style_id]
    if int(source_xf.attrib.get("fontId", "0")) == font_id:
        source_xf.set("applyFont", "1")
        return source_style_id

    clone = deepcopy(source_xf)
    clone.set("fontId", str(font_id))
    clone.set("applyFont", "1")
    cell_xfs.append(clone)
    xfs.append(clone)
    cell_xfs.set("count", str(len(xfs)))
    return len(xfs) - 1


def _clone_cell_xf_with_num_fmt(
    cell_xfs,
    xfs,
    *,
    source_style_id: int,
    num_fmt_id: int,
) -> int:
    source_xf = xfs[source_style_id]
    if int(source_xf.attrib.get("numFmtId", "0")) == num_fmt_id:
        source_xf.set("applyNumberFormat", "1")
        return source_style_id

    clone = deepcopy(source_xf)
    clone.set("numFmtId", str(num_fmt_id))
    clone.set("applyNumberFormat", "1")
    cell_xfs.append(clone)
    xfs.append(clone)
    cell_xfs.set("count", str(len(xfs)))
    return len(xfs) - 1


def _clone_cell_xf_with_center_alignment(
    cell_xfs,
    xfs,
    *,
    source_style_id: int,
) -> int:
    source_xf = xfs[source_style_id]
    alignment = source_xf.find(f"{{{_NS_MAIN}}}alignment")
    if alignment is not None and alignment.attrib.get("horizontal") == "center":
        source_xf.set("applyAlignment", "1")
        return source_style_id

    clone = deepcopy(source_xf)
    clone_alignment = clone.find(f"{{{_NS_MAIN}}}alignment")
    if clone_alignment is None:
        clone_alignment = LET.Element(f"{{{_NS_MAIN}}}alignment")
        clone.append(clone_alignment)
    clone_alignment.set("horizontal", "center")
    clone_alignment.set("vertical", "center")
    clone.set("applyAlignment", "1")
    cell_xfs.append(clone)
    xfs.append(clone)
    cell_xfs.set("count", str(len(xfs)))
    return len(xfs) - 1


def _clone_cell_xf_with_right_alignment(
    cell_xfs,
    xfs,
    *,
    source_style_id: int,
) -> int:
    source_xf = xfs[source_style_id]
    alignment = source_xf.find(f"{{{_NS_MAIN}}}alignment")
    if alignment is not None and alignment.attrib.get("horizontal") == "right":
        source_xf.set("applyAlignment", "1")
        return source_style_id

    clone = deepcopy(source_xf)
    clone_alignment = clone.find(f"{{{_NS_MAIN}}}alignment")
    if clone_alignment is None:
        clone_alignment = LET.Element(f"{{{_NS_MAIN}}}alignment")
        clone.append(clone_alignment)
    clone_alignment.set("horizontal", "right")
    clone_alignment.set("vertical", "center")
    clone.set("applyAlignment", "1")
    cell_xfs.append(clone)
    xfs.append(clone)
    cell_xfs.set("count", str(len(xfs)))
    return len(xfs) - 1


def _clone_cell_xf_with_border_id(
    cell_xfs,
    xfs,
    *,
    source_style_id: int,
    border_id: int,
) -> int:
    source_xf = xfs[source_style_id]
    if int(source_xf.attrib.get("borderId", "0")) == border_id:
        source_xf.set("applyBorder", "1")
        return source_style_id

    clone = deepcopy(source_xf)
    clone.set("borderId", str(border_id))
    clone.set("applyBorder", "1")
    cell_xfs.append(clone)
    xfs.append(clone)
    cell_xfs.set("count", str(len(xfs)))
    return len(xfs) - 1


def _patch_styles_xml(
    template_styles_xml: bytes,
    template_cell_styles_by_sheet: dict[str, dict[str, str]],
) -> tuple[bytes, dict[str, dict[str, str]]]:
    ns = {"x": _NS_MAIN}
    parser = LET.XMLParser(resolve_entities=False, no_network=True, recover=False)
    styles_root = LET.fromstring(template_styles_xml, parser=parser)
    currency_num_fmt_id = 185
    millions_num_fmt_id = 186
    no_decimal_currency_num_fmt_id = 5
    no_decimal_millions_num_fmt_id = 3

    cell_xfs = styles_root.find("x:cellXfs", ns)
    if cell_xfs is None:
        return LET.tostring(styles_root, encoding="UTF-8", xml_declaration=True, standalone=True), {}

    xfs = list(cell_xfs.findall("x:xf", ns))
    if not xfs:
        return LET.tostring(styles_root, encoding="UTF-8", xml_declaration=True, standalone=True), {}

    num_fmt_code_by_id: dict[int, str] = {}
    num_fmts = styles_root.find("x:numFmts", ns)
    if num_fmts is not None:
        for num_fmt in num_fmts.findall("x:numFmt", ns):
            raw_id = num_fmt.attrib.get("numFmtId")
            if raw_id is None:
                continue
            try:
                num_fmt_code_by_id[int(raw_id)] = num_fmt.attrib.get("formatCode", "")
            except ValueError:
                continue

    def _ensure_custom_num_fmt(format_code: str, preferred_id: int = 190) -> int:
        nonlocal num_fmts
        for fmt_id, code in num_fmt_code_by_id.items():
            if code == format_code:
                return fmt_id

        if num_fmts is None:
            num_fmts = LET.Element(f"{{{_NS_MAIN}}}numFmts")
            num_fmts.set("count", "0")
            insert_at = 0
            if len(styles_root) > 0 and styles_root[0].tag.endswith("numFmts"):
                styles_root.remove(styles_root[0])
            if len(styles_root) > 0 and styles_root[0].tag.endswith("fonts"):
                styles_root.insert(0, num_fmts)
            else:
                styles_root.insert(insert_at, num_fmts)

        used_ids = set(num_fmt_code_by_id.keys())
        fmt_id = preferred_id
        while fmt_id in used_ids:
            fmt_id += 1

        node = LET.SubElement(num_fmts, f"{{{_NS_MAIN}}}numFmt")
        node.set("numFmtId", str(fmt_id))
        node.set("formatCode", format_code)
        num_fmt_code_by_id[fmt_id] = format_code
        num_fmts.set("count", str(len(num_fmts.findall("x:numFmt", ns))))
        return fmt_id

    # Avoid locale-dependent built-in currency symbols (e.g., AED) in outputs.
    no_decimal_currency_num_fmt_id = _ensure_custom_num_fmt('"$"#,##0_);("$"#,##0);-')

    def _is_percent_style(style_id: int) -> bool:
        if not (0 <= style_id < len(xfs)):
            return False
        num_fmt_id = int(xfs[style_id].attrib.get("numFmtId", "0"))
        if num_fmt_id in {9, 10}:
            return True
        return "%" in num_fmt_code_by_id.get(num_fmt_id, "")

    dcf_base_styles = template_cell_styles_by_sheet.get("DCF Model - Base (1)", {})
    outputs_styles = template_cell_styles_by_sheet.get("Outputs - Base") or template_cell_styles_by_sheet.get("Ouputs - Base", {})
    cover_styles = template_cell_styles_by_sheet.get("Cover", {})
    wacc_styles = template_cell_styles_by_sheet.get("WACC", {})
    dcf_c16_style_id = int(dcf_base_styles.get("C16", "-1"))
    blue_font_anchor_style_id = int(dcf_base_styles.get("F9", "-1"))
    black_font_anchor_style_id = int(dcf_base_styles.get("B20", "0"))
    currency_style_ids = {
        int(dcf_base_styles.get("H20", "-1")),
        int(dcf_base_styles.get("Q94", "-1")),
    }
    for style_id in currency_style_ids:
        if not (0 <= style_id < len(xfs)):
            continue
        xf = xfs[style_id]
        xf.set("numFmtId", str(currency_num_fmt_id))
        xf.set("applyNumberFormat", "1")

    year_style_ids = {
        int(outputs_styles.get("H6", "-1")),
        int((template_cell_styles_by_sheet.get("Data Given (Recalculated)", {})).get("G6", "-1")),
    }
    for style_id in year_style_ids:
        if not (0 <= style_id < len(xfs)):
            continue
        xf = xfs[style_id]
        xf.set("numFmtId", "49")
        xf.set("applyNumberFormat", "1")

    if not (0 <= blue_font_anchor_style_id < len(xfs)):
        return LET.tostring(styles_root, encoding="UTF-8", xml_declaration=True, standalone=True), {}

    blue_font_id = int(xfs[blue_font_anchor_style_id].attrib.get("fontId", "0"))
    black_font_id = 0
    if 0 <= black_font_anchor_style_id < len(xfs):
        black_font_id = int(xfs[black_font_anchor_style_id].attrib.get("fontId", "0"))
    if not (0 <= dcf_c16_style_id < len(xfs)):
        return LET.tostring(styles_root, encoding="UTF-8", xml_declaration=True, standalone=True), {}

    dcf_c16_blue_style_id = _clone_cell_xf_with_font(
        cell_xfs,
        xfs,
        source_style_id=dcf_c16_style_id,
        font_id=blue_font_id,
    )
    dcf_sheet_names = _DCF_SCENARIO_SHEET_NAMES
    style_id_overrides_by_sheet = _initialize_dcf_style_overrides("C16", dcf_c16_blue_style_id)
    style_id_overrides_by_sheet["Cover"] = {}

    currency_clone_by_source: dict[int, int] = {}

    def _currency_style_for(source_style_id: int) -> int:
        cached = currency_clone_by_source.get(source_style_id)
        if cached is not None:
            return cached
        style_id = _clone_cell_xf_with_num_fmt(
            cell_xfs,
            xfs,
            source_style_id=source_style_id,
            num_fmt_id=currency_num_fmt_id,
        )
        currency_clone_by_source[source_style_id] = style_id
        return style_id

    no_decimal_currency_clone_by_source: dict[int, int] = {}

    def _no_decimal_currency_style_for(source_style_id: int) -> int:
        cached = no_decimal_currency_clone_by_source.get(source_style_id)
        if cached is not None:
            return cached
        style_id = _clone_cell_xf_with_num_fmt(
            cell_xfs,
            xfs,
            source_style_id=source_style_id,
            num_fmt_id=no_decimal_currency_num_fmt_id,
        )
        no_decimal_currency_clone_by_source[source_style_id] = style_id
        return style_id

    for address in ("C14", "C15"):
        source_style_raw = cover_styles.get(address)
        if source_style_raw is None:
            continue
        source_style_id = int(source_style_raw)
        if not (0 <= source_style_id < len(xfs)):
            continue
        cover_style_id = _no_decimal_currency_style_for(source_style_id)
        style_id_overrides_by_sheet["Cover"][address] = str(cover_style_id)

    centered_clone_by_source: dict[int, int] = {}
    right_aligned_clone_by_source: dict[int, int] = {}

    def _centered_style_for(source_style_id: int) -> int:
        cached = centered_clone_by_source.get(source_style_id)
        if cached is not None:
            return cached
        style_id = _clone_cell_xf_with_center_alignment(
            cell_xfs,
            xfs,
            source_style_id=source_style_id,
        )
        centered_clone_by_source[source_style_id] = style_id
        return style_id

    def _right_aligned_style_for(source_style_id: int) -> int:
        cached = right_aligned_clone_by_source.get(source_style_id)
        if cached is not None:
            return cached
        style_id = _clone_cell_xf_with_right_alignment(
            cell_xfs,
            xfs,
            source_style_id=source_style_id,
        )
        right_aligned_clone_by_source[source_style_id] = style_id
        return style_id

    border_clone_by_source_and_border: dict[tuple[int, int], int] = {}

    def _style_with_border_from(source_style_id: int, border_style_id: int) -> int:
        border_id = int(xfs[border_style_id].attrib.get("borderId", "0"))
        key = (source_style_id, border_id)
        cached = border_clone_by_source_and_border.get(key)
        if cached is not None:
            return cached
        style_id = _clone_cell_xf_with_border_id(
            cell_xfs,
            xfs,
            source_style_id=source_style_id,
            border_id=border_id,
        )
        border_clone_by_source_and_border[key] = style_id
        return style_id

    currency_rows = (
        list(range(24, 33))
        + list(range(34, 47))
        + list(range(48, 58))
        + list(range(60, 78))
        + list(range(85, 116))
    )
    currency_targets = {"C9"}
    for row in currency_rows:
        for col in "JKLMNOPQ":
            currency_targets.add(f"{col}{row}")

    for address in currency_targets:
        source_style_raw = dcf_base_styles.get(address)
        if source_style_raw is None:
            continue
        source_style_id = int(source_style_raw)
        if not (0 <= source_style_id < len(xfs)):
            continue
        if _is_percent_style(source_style_id):
            continue
        currency_style_id = _currency_style_for(source_style_id)
        _apply_override_to_sheets(style_id_overrides_by_sheet, dcf_sheet_names, address, currency_style_id)

    millions_clone_by_source: dict[int, int] = {}

    def _millions_style_for(source_style_id: int) -> int:
        cached = millions_clone_by_source.get(source_style_id)
        if cached is not None:
            return cached
        style_id = _clone_cell_xf_with_num_fmt(
            cell_xfs,
            xfs,
            source_style_id=source_style_id,
            num_fmt_id=millions_num_fmt_id,
        )
        millions_clone_by_source[source_style_id] = style_id
        return style_id

    percent_source_style_raw = dcf_base_styles.get("F11")
    percent_num_fmt_id = 171
    if percent_source_style_raw is not None:
        percent_source_style_id = int(percent_source_style_raw)
        if 0 <= percent_source_style_id < len(xfs):
            percent_num_fmt_id = int(xfs[percent_source_style_id].attrib.get("numFmtId", "171"))

    percent_clone_by_source: dict[int, int] = {}

    def _percent_style_for(source_style_id: int) -> int:
        cached = percent_clone_by_source.get(source_style_id)
        if cached is not None:
            return cached
        style_id = _clone_cell_xf_with_num_fmt(
            cell_xfs,
            xfs,
            source_style_id=source_style_id,
            num_fmt_id=percent_num_fmt_id,
        )
        percent_clone_by_source[source_style_id] = style_id
        return style_id

    black_font_clone_by_source: dict[int, int] = {}

    def _black_font_style_for(source_style_id: int) -> int:
        cached = black_font_clone_by_source.get(source_style_id)
        if cached is not None:
            return cached
        style_id = _clone_cell_xf_with_font(
            cell_xfs,
            xfs,
            source_style_id=source_style_id,
            font_id=black_font_id,
        )
        black_font_clone_by_source[source_style_id] = style_id
        return style_id

    blue_font_clone_by_source: dict[int, int] = {}

    def _blue_font_style_for(source_style_id: int) -> int:
        cached = blue_font_clone_by_source.get(source_style_id)
        if cached is not None:
            return cached
        style_id = _clone_cell_xf_with_font(
            cell_xfs,
            xfs,
            source_style_id=source_style_id,
            font_id=blue_font_id,
        )
        blue_font_clone_by_source[source_style_id] = style_id
        return style_id

    sensitivity_millions_black_clone_by_source: dict[int, int] = {}

    def _sensitivity_millions_black_style_for(source_style_id: int) -> int:
        cached = sensitivity_millions_black_clone_by_source.get(source_style_id)
        if cached is not None:
            return cached
        num_fmt_style_id = _clone_cell_xf_with_num_fmt(
            cell_xfs,
            xfs,
            source_style_id=source_style_id,
            num_fmt_id=no_decimal_millions_num_fmt_id,
        )
        style_id = _clone_cell_xf_with_font(
            cell_xfs,
            xfs,
            source_style_id=num_fmt_style_id,
            font_id=black_font_id,
        )
        sensitivity_millions_black_clone_by_source[source_style_id] = style_id
        return style_id

    statement_amount_rows = {
        20, 24, 27, 30, 32, 36, 39, 42, 45, 48, 51, 54, 57, 60, 65, 68, 74, 75, 76, 77, 78
    }
    statement_percent_rows = {
        21, 25, 28, 33, 37, 40, 43, 46, 49, 52, 55, 58, 61, 66, 69, 70, 79, 85
    }

    for row in statement_amount_rows:
        for col in "GHIJKLMNOPQ":
            address = f"{col}{row}"
            source_style_raw = dcf_base_styles.get(address)
            if source_style_raw is None:
                continue
            source_style_id = int(source_style_raw)
            if not (0 <= source_style_id < len(xfs)):
                continue
            millions_style_id = _no_decimal_currency_style_for(source_style_id)
            _apply_override_to_sheets(style_id_overrides_by_sheet, dcf_sheet_names, address, millions_style_id)

    for row in statement_percent_rows:
        for col in "GHIJKLMNOPQ":
            address = f"{col}{row}"
            source_style_raw = dcf_base_styles.get(address)
            if source_style_raw is None:
                continue
            source_style_id = int(source_style_raw)
            if not (0 <= source_style_id < len(xfs)):
                continue
            percent_style_id = _percent_style_for(source_style_id)
            _apply_override_to_sheets(style_id_overrides_by_sheet, dcf_sheet_names, address, percent_style_id)

    # Template styles for some percent rows can be inconsistent on historical
    # columns (H/I). Force these rows to percent style across all timeline cols.
    percent_source_fallback = dcf_base_styles.get("F11")
    if percent_source_fallback is not None:
        percent_source_id = int(percent_source_fallback)
        if 0 <= percent_source_id < len(xfs):
            forced_percent_style_id = _percent_style_for(percent_source_id)
            for row in (21, 85):
                for col in "GHIJKLMNOPQ":
                    address = f"{col}{row}"
                    _apply_override_to_sheets(style_id_overrides_by_sheet, dcf_sheet_names, address, forced_percent_style_id)

    sensitivity_percent_style_id: int | None = None
    sensitivity_percent_source = dcf_base_styles.get("C120")
    if sensitivity_percent_source is not None:
        sensitivity_percent_source_id = int(sensitivity_percent_source)
        if 0 <= sensitivity_percent_source_id < len(xfs):
            sensitivity_percent_style_id = _percent_style_for(sensitivity_percent_source_id)
            for address in (
                "D119", "E119", "F119", "G119", "H119",
                "C120", "C121", "C122", "C123", "C124",
                "J119", "K119", "L119", "M119", "N119",
                "I120", "I121", "I122", "I123", "I124",
            ):
                style_id_overrides_by_sheet["DCF Model - Base (1)"][address] = str(sensitivity_percent_style_id)

    # Explicit assumption/formula font semantics (applied late so they survive
    # currency/percent format override passes above):
    # - C9 is formula-driven (black)
    # - F13 is hardcoded input (blue)
    explicit_font_overrides = {
        "C9": _black_font_style_for,
        "F13": _blue_font_style_for,
    }
    for sheet_name in dcf_sheet_names:
        for address, style_fn in explicit_font_overrides.items():
            source_style_raw = style_id_overrides_by_sheet[sheet_name].get(address, dcf_base_styles.get(address))
            if source_style_raw is None:
                continue
            source_style_id = int(source_style_raw)
            if not (0 <= source_style_id < len(xfs)):
                continue
            font_style_id = style_fn(source_style_id)
            style_id_overrides_by_sheet[sheet_name][address] = str(font_style_id)

    outputs_sheet_name = "Outputs - Base"
    if outputs_sheet_name not in style_id_overrides_by_sheet:
        style_id_overrides_by_sheet[outputs_sheet_name] = {}

    outputs_amount_rows = set(range(8, 17)) | set(range(23, 43))
    for row in outputs_amount_rows:
        for col in ("D", *tuple("HIJKLMNOPQ")):
            address = f"{col}{row}"
            source_style_raw = outputs_styles.get(address)
            if source_style_raw is None:
                continue
            source_style_id = int(source_style_raw)
            if not (0 <= source_style_id < len(xfs)):
                continue
            if _is_percent_style(source_style_id):
                continue
            no_decimal_style_id = _no_decimal_currency_style_for(source_style_id)
            style_id_overrides_by_sheet[outputs_sheet_name][address] = str(no_decimal_style_id)

    # Force EBITDA row on Outputs to whole-number currency (no decimals).
    for col in "HIJKLMNOPQ":
        address = f"{col}10"
        source_style_raw = outputs_styles.get(address)
        if source_style_raw is None:
            source_style_raw = outputs_styles.get("H10")
        if source_style_raw is None:
            continue
        source_style_id = int(source_style_raw)
        if not (0 <= source_style_id < len(xfs)):
            continue
        no_decimal_style_id = _no_decimal_currency_style_for(source_style_id)
        style_id_overrides_by_sheet[outputs_sheet_name][address] = str(no_decimal_style_id)

    # Force discount rate row on Outputs to percentage formatting.
    percent_source_raw = outputs_styles.get("J22") or outputs_styles.get("D27")
    if percent_source_raw is not None:
        percent_source_id = int(percent_source_raw)
        if 0 <= percent_source_id < len(xfs):
            outputs_percent_style_id = _percent_style_for(percent_source_id)
            for col in "HIJKLMNOPQ":
                style_id_overrides_by_sheet[outputs_sheet_name][f"{col}22"] = str(outputs_percent_style_id)

    sensitivity_amount_cells = {
        *(f"{col}{row}" for row in range(120, 125) for col in "DEFGH"),
        *(f"{col}{row}" for row in range(120, 125) for col in "JKLMN"),
        *(f"C{row}" for row in range(126, 131)),
        *(f"J{row}" for row in range(126, 131)),
    }
    for address in sensitivity_amount_cells:
        source_style_raw = dcf_base_styles.get(address)
        if source_style_raw is None:
            match = re.fullmatch(r"([A-Z]+)(\d+)", address)
            if match:
                col_letters, row_text = match.groups()
                row_num = int(row_text)
                if col_letters in {"J", "K", "L", "M", "N"} and 120 <= row_num <= 124:
                    source_style_raw = dcf_base_styles.get(f"{_column_number_to_letters(ord(col_letters) - ord('J') + ord('D') - ord('A') + 1)}{row_num}")
                elif col_letters == "J" and 126 <= row_num <= 130:
                    source_style_raw = dcf_base_styles.get(f"C{row_num}")
        if source_style_raw is None:
            continue
        source_style_id = int(source_style_raw)
        if not (0 <= source_style_id < len(xfs)):
            continue
        millions_style_id = _sensitivity_millions_black_style_for(source_style_id)
        _apply_override_to_sheets(style_id_overrides_by_sheet, dcf_sheet_names, address, millions_style_id)

    sensitivity_header_number_cells = {
        *(f"{col}119" for col in "DEFGH"),
        *(f"{col}119" for col in "JKLMN"),
        *(f"I{row}" for row in range(120, 125)),
        *(f"C{row}" for row in range(120, 125)),
    }
    left_sensitivity_header_cells = {f"{col}119" for col in "DEFGH"}
    for address in sensitivity_header_number_cells:
        if address in left_sensitivity_header_cells and sensitivity_percent_style_id is not None:
            black_percent_style_id = _black_font_style_for(sensitivity_percent_style_id)
            _apply_override_to_sheets(style_id_overrides_by_sheet, dcf_sheet_names, address, black_percent_style_id)
            continue
        source_style_raw = dcf_base_styles.get(address)
        if source_style_raw is None:
            continue
        source_style_id = int(source_style_raw)
        if not (0 <= source_style_id < len(xfs)):
            continue
        black_style_id = _black_font_style_for(source_style_id)
        _apply_override_to_sheets(style_id_overrides_by_sheet, dcf_sheet_names, address, black_style_id)

    # Keep sensitivity summary table borders consistent in Bull/Bear by
    # forcing the same border-bearing styles as Base.
    sensitivity_border_cells = {
        *(f"{col}{row}" for col in "BC" for row in range(126, 131)),
        *(f"{col}{row}" for col in "I" for row in range(126, 131)),
    }
    for address in sensitivity_border_cells:
        source_style_raw = dcf_base_styles.get(address)
        match = re.fullmatch(r"([A-Z]+)(\d+)", address)
        if match is None:
            continue
        col_letters, row_text = match.groups()
        row_num = int(row_text)

        # Left table cells can directly inherit base styles.
        if col_letters in {"B", "C"}:
            if source_style_raw is None:
                alt_col = "C" if col_letters == "B" else "B"
                source_style_raw = dcf_base_styles.get(f"{alt_col}{row_num}")
                if source_style_raw is None:
                    continue
            _apply_override_to_sheets(style_id_overrides_by_sheet, dcf_sheet_names, address, source_style_raw)
            continue

        # Right table: keep cell's own number format/font and only borrow border.
        target_style_raw = dcf_base_styles.get(address)
        if target_style_raw is None:
            target_style_raw = dcf_base_styles.get(f"B{row_num}")
        if target_style_raw is None:
            continue
        border_source_raw = dcf_base_styles.get(f"B{row_num}")
        if border_source_raw is None:
            continue
        target_style_id = int(target_style_raw)
        border_source_id = int(border_source_raw)
        if not (0 <= target_style_id < len(xfs) and 0 <= border_source_id < len(xfs)):
            continue
        merged_style_id = _style_with_border_from(target_style_id, border_source_id)
        _apply_override_to_sheets(style_id_overrides_by_sheet, dcf_sheet_names, address, merged_style_id)

    # Right-align timeline year headers in core DCF sections across all scenarios.
    right_aligned_header_cells = {
        *(f"{col}{row}" for row in (18, 63, 72, 89) for col in "HIJKLMNOPQ"),
        *(f"G{row}" for row in (18, 63, 72)),
    }
    for address in right_aligned_header_cells:
        source_style_raw = dcf_base_styles.get(address)
        if source_style_raw is None:
            source_style_raw = dcf_base_styles.get("H18")
        if source_style_raw is None:
            continue
        source_style_id = int(source_style_raw)
        if not (0 <= source_style_id < len(xfs)):
            continue
        right_style_id = _right_aligned_style_for(source_style_id)
        text_right_style_id = _clone_cell_xf_with_num_fmt(
            cell_xfs,
            xfs,
            source_style_id=right_style_id,
            num_fmt_id=49,
        )
        _apply_override_to_sheets(style_id_overrides_by_sheet, dcf_sheet_names, address, text_right_style_id)

    # Right-align FY headers everywhere they appear (Outputs + data sheets).
    outputs_right_aligned_cells = {f"{col}6" for col in "HIJKLMNOPQ"}
    if outputs_sheet_name not in style_id_overrides_by_sheet:
        style_id_overrides_by_sheet[outputs_sheet_name] = {}
    for address in outputs_right_aligned_cells:
        source_style_raw = outputs_styles.get(address)
        if source_style_raw is None:
            source_style_raw = outputs_styles.get("H6")
        if source_style_raw is None:
            continue
        source_style_id = int(source_style_raw)
        if not (0 <= source_style_id < len(xfs)):
            continue
        right_style_id = _right_aligned_style_for(source_style_id)
        # Force textual FY labels even when locale/template fallback styles differ.
        text_right_style_id = _clone_cell_xf_with_num_fmt(
            cell_xfs,
            xfs,
            source_style_id=right_style_id,
            num_fmt_id=49,
        )
        style_id_overrides_by_sheet[outputs_sheet_name][address] = str(text_right_style_id)

    data_recalc_name = "Data Given (Recalculated)"
    data_recalc_styles = template_cell_styles_by_sheet.get(data_recalc_name, {})
    if data_recalc_name not in style_id_overrides_by_sheet:
        style_id_overrides_by_sheet[data_recalc_name] = {}
    for address in {f"{col}6" for col in "GHIJKLMNOP"}:
        source_style_raw = data_recalc_styles.get(address)
        if source_style_raw is None:
            continue
        source_style_id = int(source_style_raw)
        if not (0 <= source_style_id < len(xfs)):
            continue
        style_id_overrides_by_sheet[data_recalc_name][address] = str(_right_aligned_style_for(source_style_id))

    data_original_name = "Original & Adjusted Data"
    data_original_styles = template_cell_styles_by_sheet.get(data_original_name, {})
    if data_original_name not in style_id_overrides_by_sheet:
        style_id_overrides_by_sheet[data_original_name] = {}
    for address in {f"{col}6" for col in ("V", "W", "X", "Y", "Z", "AA", "AB", "AC", "AD", "AE")}:
        source_style_raw = data_original_styles.get(address)
        if source_style_raw is None:
            continue
        source_style_id = int(source_style_raw)
        if not (0 <= source_style_id < len(xfs)):
            continue
        style_id_overrides_by_sheet[data_original_name][address] = str(_right_aligned_style_for(source_style_id))

    # Display present value cash flow line as whole currency amounts (no cents).
    for col in "HIJKLMNOPQ":
        address = f"{col}87"
        source_style_raw = dcf_base_styles.get(address)
        if source_style_raw is None:
            continue
        source_style_id = int(source_style_raw)
        if not (0 <= source_style_id < len(xfs)):
            continue
        no_decimal_style_id = _no_decimal_currency_style_for(source_style_id)
        _apply_override_to_sheets(style_id_overrides_by_sheet, dcf_sheet_names, address, no_decimal_style_id)

    # Keep prior-year (G) headers visually identical to the first timeline
    # header cells so FY2020A doesn't lose bold/heading emphasis.
    for sheet_name in dcf_sheet_names:
        for g_cell, h_cell in (("G18", "H18"), ("G63", "H63"), ("G72", "H72")):
            header_style = style_id_overrides_by_sheet[sheet_name].get(h_cell)
            if header_style is None:
                header_style = dcf_base_styles.get(h_cell)
            if header_style is not None:
                style_id_overrides_by_sheet[sheet_name][g_cell] = header_style

    # Keep WACC peer equity values consistently formatted in millions.
    wacc_equity_cells = {
        *(f"J{row}" for row in range(8, 14)),
        "J16",
        "J17",
        "J18",
        "J19",
        "J23",
    }
    if "WACC" not in style_id_overrides_by_sheet:
        style_id_overrides_by_sheet["WACC"] = {}
    for address in wacc_equity_cells:
        source_style_raw = wacc_styles.get(address)
        if source_style_raw is None:
            continue
        source_style_id = int(source_style_raw)
        if not (0 <= source_style_id < len(xfs)):
            continue
        millions_style_id = _millions_style_for(source_style_id)
        style_id_overrides_by_sheet["WACC"][address] = str(millions_style_id)

    return (
        LET.tostring(styles_root, encoding="UTF-8", xml_declaration=True, standalone=True),
        style_id_overrides_by_sheet,
    )


def _patch_sheet_style_ids(
    sheet_xml: bytes,
    template_sheet_xml: bytes,
    sheet_name: str | None = None,
    style_id_overrides: dict[str, str] | None = None,
) -> bytes:
    ns = {"x": _NS_MAIN}

    source_root = ET.fromstring(sheet_xml)
    template_root = ET.fromstring(template_sheet_xml)

    template_cell_styles = {}
    for cell in template_root.findall(".//x:c", ns):
        address = cell.attrib.get("r")
        if not address:
            continue
        template_cell_styles[address] = cell.attrib.get("s")

    template_row_styles = {}
    for row in template_root.findall(".//x:row", ns):
        row_num = row.attrib.get("r")
        if not row_num:
            continue
        template_row_styles[row_num] = row.attrib.get("s")

    template_col_styles = {}
    for col in template_root.findall(".//x:col", ns):
        min_col = col.attrib.get("min")
        max_col = col.attrib.get("max")
        if not min_col or not max_col:
            continue
        template_col_styles[(min_col, max_col)] = col.attrib.get("style")

    for cell in source_root.findall(".//x:c", ns):
        address = cell.attrib.get("r")
        if not address:
            continue
        forced_style_id = style_id_overrides.get(address) if style_id_overrides else None
        if forced_style_id is not None:
            cell.set("s", forced_style_id)
            continue

        style_override_source = _style_override_source_address(sheet_name, address)
        if style_override_source:
            override_style_id = template_cell_styles.get(style_override_source)
            if override_style_id is None:
                cell.attrib.pop("s", None)
            else:
                cell.set("s", override_style_id)
            continue

        if address in template_cell_styles:
            style_id = template_cell_styles[address]
            if style_id is None:
                cell.attrib.pop("s", None)
            else:
                cell.set("s", style_id)
            continue

        # Fallback only when template does not explicitly define a style for the
        # right-side sensitivity matrix cell.
        mirror_address = _mirror_right_sensitivity_to_left(address)
        if mirror_address:
            mirror_style_id = template_cell_styles.get(mirror_address)
            if mirror_style_id is None:
                cell.attrib.pop("s", None)
            else:
                cell.set("s", mirror_style_id)
            continue

        continue

    for row in source_root.findall(".//x:row", ns):
        row_num = row.attrib.get("r")
        if not row_num or row_num not in template_row_styles:
            continue
        style_id = template_row_styles[row_num]
        if style_id is None:
            row.attrib.pop("s", None)
        else:
            row.set("s", style_id)

    for col in source_root.findall(".//x:col", ns):
        min_col = col.attrib.get("min")
        max_col = col.attrib.get("max")
        if not min_col or not max_col:
            continue
        style_id = template_col_styles.get((min_col, max_col))
        if style_id is None:
            col.attrib.pop("style", None)
        else:
            col.set("style", style_id)

    return ET.tostring(source_root, encoding="utf-8", xml_declaration=True)


def _mirror_right_sensitivity_to_left(address: str) -> str | None:
    match = re.fullmatch(r"([A-Z]+)(\d+)", address)
    if not match:
        return None

    letters, row_text = match.groups()
    row = int(row_text)
    if row < 117 or row > 130:
        return None

    col = 0
    for char in letters:
        col = (col * 26) + (ord(char) - ord("A") + 1)
    if col < 9 or col > 14:  # I..N
        return None

    right_to_left = {
        9: 2,   # I -> B
        10: 4,  # J -> D
        11: 5,  # K -> E
        12: 6,  # L -> F
        13: 7,  # M -> G
        14: 8,  # N -> H
    }
    left_col = right_to_left.get(col)
    if left_col is None:
        return None
    return f"{_column_number_to_letters(left_col)}{row}"


def _style_override_source_address(sheet_name: str | None, address: str) -> str | None:
    if not sheet_name:
        return None
    sheet_overrides = _STYLE_SOURCE_OVERRIDES.get(sheet_name)
    if not sheet_overrides:
        return None
    return sheet_overrides.get(address)


def _column_number_to_letters(col_num: int) -> str:
    letters = ""
    while col_num > 0:
        col_num, rem = divmod(col_num - 1, 26)
        letters = chr(ord("A") + rem) + letters
    return letters


def _initialize_dcf_style_overrides(initial_address: str, style_id: int) -> dict[str, dict[str, str]]:
    return {
        sheet_name: {initial_address: str(style_id)}
        for sheet_name in _DCF_SCENARIO_SHEET_NAMES
    }


def _apply_override_to_sheets(
    style_id_overrides_by_sheet: dict[str, dict[str, str]],
    sheet_names: tuple[str, ...],
    address: str,
    style_id: int | str,
) -> None:
    style_value = str(style_id)
    for sheet_name in sheet_names:
        style_id_overrides_by_sheet[sheet_name][address] = style_value
